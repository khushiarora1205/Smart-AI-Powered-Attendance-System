from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity, get_jwt
from pymongo import MongoClient
from bson import ObjectId
import datetime
import os
import numpy as np
import base64
import cv2
from deepface import DeepFace
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import hashlib
import random
import secrets
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
CORS(app, origins=["http://localhost:5173", "http://localhost:5174", "http://localhost:3000", "http://127.0.0.1:5173", "http://127.0.0.1:5174", "http://localhost:5001", "http://127.0.0.1:5001"])

# JWT Configuration
app.config['JWT_SECRET_KEY'] = 'your-secret-key-change-in-production'  # Change this in production!
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = datetime.timedelta(hours=24)  # Token expires in 24 hours
jwt = JWTManager(app)

# JWT Error handlers
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    print("DEBUG: JWT token has expired")
    return jsonify({"error": "Token has expired", "message": "Please login again"}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error):
    print(f"DEBUG: Invalid JWT token: {error}")
    return jsonify({"error": "Invalid token", "message": "Please login again"}), 401

@jwt.unauthorized_loader
def missing_token_callback(error):
    print(f"DEBUG: Missing JWT token: {error}")
    return jsonify({"error": "Missing token", "message": "Authorization token required"}), 401

# MongoDB connection
MONGO_URI = ""
client = MongoClient(MONGO_URI)
db = client['attendance_system']
students_col = db['students']
attendance_col = db['attendance']
lectures_col = db['lectures']
users_col = db['users']  # New collection for user authentication
teachers_col = db['teachers']  # New collection for teachers managed by admin
teacher_assignments_col = db['teacher_assignments']  # Collection for teacher-subject-semester assignments
subjects_col = db['subjects']  # Collection for subjects by department and semester
mentor_assignments_col = db['mentor_assignments']  # Collection for batch+group mentor assignments
medical_leaves_col = db['medical_leaves']  # Collection for medical leave requests

# Create uploads directory for ML proofs
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'ml_proofs')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max file size

# Current lecture state
current_lecture = None

# Username generation function for teachers
def generate_unique_username(name):
    """
    Generate a unique username in format: First letter (capitalized) + 4-digit random number
    """
    if not name:
        raise ValueError("Name cannot be empty")
    
    # Get first letter and capitalize it
    first_letter = name.strip()[0].upper()
    
    # Keep generating until we find a unique username
    max_attempts = 100  # Prevent infinite loop
    attempts = 0
    
    while attempts < max_attempts:
        # Generate random 4-digit number (1000-9999)
        random_number = random.randint(1000, 9999)
        username = f"{first_letter}{random_number}"
        
        # Check if username already exists in database
        if not teachers_col.find_one({"username": username}):
            return username
        
        attempts += 1
    
    # If we couldn't find a unique username after max attempts
    raise Exception("Could not generate unique username after maximum attempts")

def generate_unique_password():
    """
    Generate a unique 8-digit numeric password
    Example: '94821653'
    """
    max_attempts = 100  # Prevent infinite loop
    attempts = 0
    
    while attempts < max_attempts:
        # Generate random 8-digit number (10000000-99999999)
        password = str(random.randint(10000000, 99999999))
        
        # Check if password already exists in database (check hashed version)
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        if not teachers_col.find_one({"password_hash": password_hash}):
            return password
        
        attempts += 1
    
    # If we couldn't find a unique password after max attempts
    raise Exception("Could not generate unique password after maximum attempts")

def send_teacher_credentials_email(teacher_name, username, password, email, department):
    """
    Send teacher credentials via email
    """
    try:
        # QUICK WORKING EMAIL SOLUTION - Using Gmail with Less Secure Apps
        # Note: This is for testing only. For production, use App Passwords.
        
        sender_email = ""
        sender_password = ""
        smtp_server = "smtp.gmail.com"
        smtp_port = 
        
        # Alternative: Use Outlook (often easier)
        # sender_email = "your_email@outlook.com"
        # sender_password = "your_password"
        # smtp_server = "smtp-mail.outlook.com" 
        # smtp_port = 587
        
        # Real email credentials configured
        print("📧 Using configured email credentials")
        # Email credentials configured - proceed with real sending
        print(f"📧 Sending email to: {email}")
        print(f"📤 From: {sender_email}")
        
        # Create message
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = email
        message["Subject"] = "Your Faculty Login Credentials – Smart AI Attendance System"
        
        # Email body
        body = f"""Hello {teacher_name},

Welcome to the Smart AI-Powered Attendance System!

Here are your login credentials:

Username: {username}
Password: {password}
Email: {email}
Department: {department}

Please log in using these credentials to access the teacher portal.

Best regards,
Smart Attendance System Admin
"""
        
        message.attach(MIMEText(body, "plain"))
        
        # Connect to server and send email
        print(f"📧 Attempting to send email to: {email}")
        print(f"📡 Using SMTP server: {smtp_server}:{smtp_port}")
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Enable TLS encryption
        print(f"🔐 Logging in with email: {sender_email}")
        server.login(sender_email, sender_password)
        text = message.as_string()
        server.sendmail(sender_email, email, text)
        server.quit()
        print(f"✅ Email sent successfully to: {email}")
        
        return True
    except Exception as e:
        print(f"❌ Email sending error: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        
        # Common Gmail errors and solutions
        error_str = str(e).lower()
        if "authentication failed" in error_str or "username and password not accepted" in error_str:
            print("\n🔧 GMAIL AUTHENTICATION ERROR - SOLUTIONS:")
            print("1. Enable 2-Step Verification in Gmail")
            print("2. Generate App Password: https://myaccount.google.com/apppasswords")
            print("3. Use the 16-character app password instead of regular password")
            print("4. Or enable 'Less secure app access' (less secure)")
        elif "connection" in error_str:
            print("\n🌐 CONNECTION ERROR - Check internet connection")
        
        return False

# Authentication Routes
@app.route('/api/login', methods=['POST'])
def manual_login():
    try:
        email = request.json.get('email')
        password = request.json.get('password')
        
        if not email or not password:
            return jsonify({"message": "Email and password required"}), 400
        
        # Find user in database
        user = users_col.find_one({"email": email})
        
        if not user:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Check password (you should use proper hashing in production)
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if user.get('password_hash') != password_hash:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Update last login
        users_col.update_one(
            {"email": email},
            {"$set": {"last_login": datetime.datetime.now()}}
        )
        
        # Create access token
        access_token = create_access_token(identity=email)
        
        return jsonify({
            "token": access_token,
            "user": {
                "id": str(user['_id']),
                "email": user['email'],
                "name": user['name'],
                "role": user.get('role', 'teacher')
            }
        }), 200
        
    except Exception as e:
        return jsonify({"message": "Login failed"}), 500

@app.route('/api/register', methods=['POST'])
def register():
    try:
        email = request.json.get('email')
        password = request.json.get('password')
        name = request.json.get('name')
        
        if not email or not password or not name:
            return jsonify({"message": "Email, password, and name required"}), 400
        
        # Check if user already exists
        if users_col.find_one({"email": email}):
            return jsonify({"message": "User already exists"}), 400
        
        # Hash password
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Create user
        user_data = {
            "email": email,
            "name": name,
            "password_hash": password_hash,
            "role": "teacher",
            "created_at": datetime.datetime.now(),
            "last_login": datetime.datetime.now()
        }
        
        result = users_col.insert_one(user_data)
        
        # Create access token
        access_token = create_access_token(identity=email)
        
        return jsonify({
            "token": access_token,
            "user": {
                "id": str(result.inserted_id),
                "email": email,
                "name": name,
                "role": "teacher"
            }
        }), 201
        
    except Exception as e:
        return jsonify({"message": "Registration failed"}), 500

@app.route('/api/logout', methods=['POST'])
@jwt_required()
def logout():
    return jsonify({"message": "Logged out successfully"}), 200

@app.route('/api/profile', methods=['GET'])
@jwt_required()
def get_profile():
    try:
        current_user_email = get_jwt_identity()
        user = users_col.find_one({"email": current_user_email})
        
        if not user:
            return jsonify({"message": "User not found"}), 404
        
        return jsonify({
            "user": {
                "id": str(user['_id']),
                "email": user['email'],
                "name": user['name'],
                "picture": user.get('picture', ''),
                "role": user.get('role', 'teacher'),
                "created_at": user.get('created_at'),
                "last_login": user.get('last_login')
            }
        }), 200
        
    except Exception as e:
        return jsonify({"message": "Failed to get profile"}), 500

# Teacher Authentication Routes
@app.route('/api/teacher/login', methods=['POST'])
def teacher_login():
    try:
        username = request.json.get('username')
        password = request.json.get('password')
        
        if not username or not password:
            return jsonify({"message": "Username and password required"}), 400
        
        # Find teacher in database
        teacher = teachers_col.find_one({"username": username})
        
        if not teacher:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Check password using SHA256 hashing
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if teacher.get('password_hash') != password_hash:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Update last login
        teachers_col.update_one(
            {"username": username},
            {"$set": {"last_login": datetime.datetime.now()}}
        )
        
        # Create access token with teacher role
        access_token = create_access_token(identity=username, additional_claims={"role": "teacher"})
        
        return jsonify({
            "token": access_token,
            "user": {
                "id": str(teacher['_id']),
                "username": teacher['username'],
                "name": teacher['name'],
                "email": teacher['email'],
                "department": teacher['department'],
                "role": "teacher"
            }
        }), 200
        
    except Exception as e:
        print(f"Teacher login error: {e}")
        return jsonify({"message": "Login failed"}), 500

# Admin Authentication Routes
@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    try:
        email = request.json.get('email')
        password = request.json.get('password')
        
        if not email or not password:
            return jsonify({"message": "Email and password required"}), 400
        
        # Hardcoded admin credentials
        ADMIN_EMAIL = "admin@faceattend.com"
        ADMIN_PASSWORD = "admin123"
        
        if email != ADMIN_EMAIL or password != ADMIN_PASSWORD:
            return jsonify({"message": "Invalid admin credentials"}), 401
        
        # Create access token with admin role
        access_token = create_access_token(identity=email, additional_claims={"role": "admin"})
        
        return jsonify({
            "token": access_token,
            "user": {
                "email": email,
                "name": "System Administrator",
                "role": "admin"
            }
        }), 200
        
    except Exception as e:
        return jsonify({"message": "Admin login failed"}), 500

# Helper function to verify admin role
def verify_admin():
    """Verify current user is admin"""
    try:
        current_user_email = get_jwt_identity()
        if current_user_email != "admin@faceattend.com":
            return False
        return True
    except:
        return False

@app.route('/api/admin/add-teacher', methods=['POST'])
@jwt_required()
def add_teacher():
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        name = data.get('name')
        email = data.get('email')
        department = data.get('department')
        contact_number = data.get('contact_number', '')
        
        if not all([name, email, department]):
            return jsonify({"message": "Name, email, and department are required"}), 400
        
        # Check if teacher already exists
        if teachers_col.find_one({"email": email}):
            return jsonify({"message": "Teacher with this email already exists"}), 400
        
        # Generate unique username and password
        username = generate_unique_username(name)
        password = generate_unique_password()
        
        # Hash password
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Create teacher record
        teacher_data = {
            "name": name,
            "username": username,
            "email": email,
            "password_hash": password_hash,
            "department": department,
            "contact_number": contact_number,
            "created_at": datetime.datetime.now()
        }
        
        result = teachers_col.insert_one(teacher_data)
        
        return jsonify({
            "success": True,
            "message": "Teacher added successfully",
            "username": username,
            "password": password,
            "email": email,
            "teacher_id": str(result.inserted_id)
        }), 201
        
    except Exception as e:
        return jsonify({"message": f"Failed to add teacher: {str(e)}"}), 500

@app.route('/api/admin/send-teacher-credentials', methods=['POST'])
@jwt_required()
def send_teacher_credentials():
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        teacher_name = data.get('name')
        username = data.get('username')
        password = data.get('password')
        email = data.get('email')
        department = data.get('department')
        
        if not all([teacher_name, username, password, email, department]):
            return jsonify({"message": "All teacher details are required"}), 400
        
        # Send email
        email_sent = send_teacher_credentials_email(teacher_name, username, password, email, department)
        
        if email_sent:
            return jsonify({
                "success": True,
                "message": f"Credentials sent successfully to {email}"
            }), 200
        else:
            return jsonify({
                "success": False,
                "message": "Failed to send email. Please try again."
            }), 500
        
    except Exception as e:
        return jsonify({"message": f"Failed to send credentials: {str(e)}"}), 500

@app.route('/api/admin/add-student', methods=['POST'])
@jwt_required()
def admin_add_student():
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        name = data.get('name')
        roll_no = data.get('roll_no')
        email = data.get('email')
        course = data.get('course')
        department = data.get('department')
        batch = data.get('batch')
        group = data.get('group')
        images = data.get('images', [])
        
        if not all([name, roll_no, email, course, department]):
            return jsonify({"message": "Name, roll number, email, course, and department are required"}), 400
        
        # Validate batch (must be 4 digits)
        if batch:
            if not batch.isdigit() or len(batch) != 4:
                return jsonify({"message": "Batch must be a 4-digit year (e.g., 2023)"}), 400
        
        # Validate group (must be non-empty if provided)
        if group is not None and not group.strip():
            return jsonify({"message": "Group cannot be empty"}), 400
        
        # Check if student already exists
        if students_col.find_one({"rollNo": roll_no}):
            return jsonify({"message": "Student with this roll number already exists"}), 400
        
        if students_col.find_one({"email": email}):
            return jsonify({"message": "Student with this email already exists"}), 400
        
        # Process face recognition if images provided
        embeddings = []
        if images:
            for idx, img_b64 in enumerate(images):
                try:
                    emb = get_embedding_from_base64(img_b64)
                    embeddings.append(emb)
                except Exception as e:
                    print(f"Failed to process image {idx+1}: {e}")
                    continue
        
        # Create student record
        student_data = {
            "name": name,
            "rollNo": roll_no,
            "email": email,
            "course": course,
            "department": department,
            "batch": batch if batch else None,
            "group": group if group else None,
            "embeddings": embeddings,
            "images": images,
            "created_at": datetime.datetime.now()
        }
        
        result = students_col.insert_one(student_data)
        
        return jsonify({
            "message": "Student added successfully",
            "student_id": str(result.inserted_id),
            "embeddings_processed": len(embeddings)
        }), 201
        
    except Exception as e:
        return jsonify({"message": f"Failed to add student: {str(e)}"}), 500

# Teacher Assignment Routes
@app.route('/api/admin/assign-subjects', methods=['POST'])
@jwt_required()
def assign_subjects_to_teacher():
    """Assign subjects and semester to a teacher"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        teacher_username = data.get('teacher_username')
        subjects = data.get('subjects', [])
        semester = data.get('semester')
        
        if not all([teacher_username, subjects, semester]):
            return jsonify({"message": "Teacher username, subjects, and semester are required"}), 400
        
        # Verify teacher exists
        teacher = teachers_col.find_one({"username": teacher_username})
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Create or update assignment
        assignment_data = {
            "teacher_username": teacher_username,
            "teacher_name": teacher.get('name'),
            "subjects": subjects,
            "semester": semester,
            "assigned_at": datetime.datetime.now(),
            "assigned_by": "admin"
        }
        
        # Use upsert to update existing assignment or create new one
        result = teacher_assignments_col.replace_one(
            {"teacher_username": teacher_username},
            assignment_data,
            upsert=True
        )
        
        return jsonify({
            "message": f"Successfully assigned {len(subjects)} subjects to {teacher.get('name')} for semester {semester}",
            "assignment": assignment_data
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to assign subjects: {str(e)}"}), 500

@app.route('/api/admin/teacher-assignments', methods=['GET'])
@jwt_required()
def get_teacher_assignments():
    """Get all teacher assignments"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        assignments = list(teacher_assignments_col.find({}, {"_id": 0}))
        return jsonify({"assignments": assignments}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to get assignments: {str(e)}"}), 500

@app.route('/api/admin/teachers', methods=['GET'])
@jwt_required()
def get_teachers_list():
    """Get list of all teachers for admin"""
    try:
        # Verify admin access  
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        teachers = list(teachers_col.find({}, {"_id": 0, "password_hash": 0}))
        return jsonify({"teachers": teachers}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to get teachers: {str(e)}"}), 500

@app.route('/api/teacher/my-assignments', methods=['GET'])
@jwt_required()
def get_my_assignments():
    """Get current teacher's subject assignments"""
    try:
        current_user = get_jwt_identity()
        
        # Find assignment for current teacher
        assignment = teacher_assignments_col.find_one(
            {"teacher_username": current_user}, 
            {"_id": 0}
        )
        
        if not assignment:
            return jsonify({
                "assignment": None,
                "message": "No subjects assigned yet"
            }), 200
        
        return jsonify({"assignment": assignment}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to get assignments: {str(e)}"}), 500

# Subjects Management Routes
@app.route('/api/subjects', methods=['GET'])
@jwt_required()
def get_subjects():
    """Get subjects filtered by department only (semester is ignored)"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        department = request.args.get('department')
        # semester parameter is ignored as subjects are now department-only
        
        # Build filter query - only filter by department
        filter_query = {}
        if department:
            filter_query['departments'] = department
        
        # Get subjects matching the criteria
        subjects = list(subjects_col.find(filter_query, {"_id": 0}))
        
        # Extract just the subject names
        subject_names = [subject['name'] for subject in subjects]
        
        return jsonify({
            "subjects": subject_names,
            "count": len(subject_names),
            "department": department
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to get subjects: {str(e)}"}), 500

@app.route('/api/admin/initialize-subjects', methods=['POST'])
@jwt_required()
def initialize_subjects():
    """Initialize subjects database with sample data (admin only)"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Check if subjects already exist
        existing_count = subjects_col.count_documents({})
        if existing_count > 0:
            return jsonify({
                "message": f"Subjects already initialized. Found {existing_count} subjects.",
                "action": "skipped"
            }), 200
        
        # Sample subjects data with department mappings (available for all semesters within department)
        sample_subjects = [
            # Computer Science Department
            {"name": "Data Structures", "departments": ["Computer Science"]},
            {"name": "Algorithms", "departments": ["Computer Science"]},
            {"name": "Database Systems", "departments": ["Computer Science"]},
            {"name": "Software Engineering", "departments": ["Computer Science"]},
            {"name": "Operating Systems", "departments": ["Computer Science"]},
            {"name": "Computer Networks", "departments": ["Computer Science"]},
            {"name": "Machine Learning", "departments": ["Computer Science"]},
            {"name": "Web Development", "departments": ["Computer Science"]},
            {"name": "Artificial Intelligence", "departments": ["Computer Science"]},
            {"name": "Computer Graphics", "departments": ["Computer Science"]},
            {"name": "Cybersecurity", "departments": ["Computer Science"]},
            {"name": "Mobile App Development", "departments": ["Computer Science"]},
            
            # Mathematics Department  
            {"name": "Calculus I", "departments": ["Mathematics"]},
            {"name": "Calculus II", "departments": ["Mathematics"]},
            {"name": "Linear Algebra", "departments": ["Mathematics"]},
            {"name": "Differential Equations", "departments": ["Mathematics"]},
            {"name": "Statistics", "departments": ["Mathematics"]},
            {"name": "Discrete Mathematics", "departments": ["Mathematics"]},
            {"name": "Probability Theory", "departments": ["Mathematics"]},
            {"name": "Number Theory", "departments": ["Mathematics"]},
            {"name": "Abstract Algebra", "departments": ["Mathematics"]},
            {"name": "Real Analysis", "departments": ["Mathematics"]},
            
            # Physics Department
            {"name": "Classical Mechanics", "departments": ["Physics"]},
            {"name": "Quantum Physics", "departments": ["Physics"]},
            {"name": "Thermodynamics", "departments": ["Physics"]},
            {"name": "Electromagnetism", "departments": ["Physics"]},
            {"name": "Optics", "departments": ["Physics"]},
            {"name": "Nuclear Physics", "departments": ["Physics"]},
            {"name": "Solid State Physics", "departments": ["Physics"]},
            {"name": "Astrophysics", "departments": ["Physics"]},
            
            # Business Department
            {"name": "Business Management", "departments": ["Business"]},
            {"name": "Marketing", "departments": ["Business"]},
            {"name": "Finance", "departments": ["Business"]},
            {"name": "Economics", "departments": ["Business"]},
            {"name": "Accounting", "departments": ["Business"]},
            {"name": "Human Resources", "departments": ["Business"]},
            {"name": "Operations Management", "departments": ["Business"]},
            {"name": "Business Analytics", "departments": ["Business"]},
            {"name": "Entrepreneurship", "departments": ["Business"]},
            
            # Engineering Department (General)
            {"name": "Engineering Physics", "departments": ["Engineering"]},
            {"name": "Circuit Analysis", "departments": ["Engineering"]},
            {"name": "Digital Logic", "departments": ["Engineering"]},
            {"name": "Control Systems", "departments": ["Engineering"]},
            {"name": "Engineering Mathematics", "departments": ["Engineering"]},
            {"name": "Engineering Drawing", "departments": ["Engineering"]},
            {"name": "Materials Science", "departments": ["Engineering"]},
            
            # ECE (Electronics and Communication Engineering) Department
            {"name": "Electronic Devices", "departments": ["ECE"]},
            {"name": "Analog Electronics", "departments": ["ECE"]},
            {"name": "Digital Electronics", "departments": ["ECE"]},
            {"name": "Communication Systems", "departments": ["ECE"]},
            {"name": "Microprocessors", "departments": ["ECE"]},
            {"name": "VLSI Design", "departments": ["ECE"]},
            {"name": "Signal Processing", "departments": ["ECE"]},
            {"name": "Embedded Systems", "departments": ["ECE"]},
            {"name": "Antenna Theory", "departments": ["ECE"]},
            {"name": "RF Engineering", "departments": ["ECE"]},
            {"name": "Power Electronics", "departments": ["ECE"]},
            {"name": "Digital Signal Processing", "departments": ["ECE"]},
            {"name": "Wireless Communication", "departments": ["ECE"]},
            {"name": "Fiber Optic Communication", "departments": ["ECE"]},
            {"name": "Satellite Communication", "departments": ["ECE"]},
            
            # Mechanical Engineering Department
            {"name": "Thermodynamics", "departments": ["Mechanical"]},
            {"name": "Fluid Mechanics", "departments": ["Mechanical"]},
            {"name": "Machine Design", "departments": ["Mechanical"]},
            {"name": "Heat Transfer", "departments": ["Mechanical"]},
            {"name": "Manufacturing Processes", "departments": ["Mechanical"]},
            {"name": "Strength of Materials", "departments": ["Mechanical"]},
            {"name": "Dynamics of Machinery", "departments": ["Mechanical"]},
            {"name": "Automobile Engineering", "departments": ["Mechanical"]},
            {"name": "Industrial Engineering", "departments": ["Mechanical"]},
            {"name": "Mechanical Vibrations", "departments": ["Mechanical"]},
            
            # Civil Engineering Department  
            {"name": "Structural Analysis", "departments": ["Civil"]},
            {"name": "Concrete Technology", "departments": ["Civil"]},
            {"name": "Geotechnical Engineering", "departments": ["Civil"]},
            {"name": "Transportation Engineering", "departments": ["Civil"]},
            {"name": "Hydraulics", "departments": ["Civil"]},
            {"name": "Environmental Engineering", "departments": ["Civil"]},
            {"name": "Construction Management", "departments": ["Civil"]},
            {"name": "Surveying", "departments": ["Civil"]},
            {"name": "Earthquake Engineering", "departments": ["Civil"]},
            {"name": "Water Resources Engineering", "departments": ["Civil"]},
            
            # Electrical Engineering Department
            {"name": "Electrical Machines", "departments": ["Electrical"]},
            {"name": "Power Systems", "departments": ["Electrical"]},
            {"name": "Electrical Circuits", "departments": ["Electrical"]},
            {"name": "Power Electronics", "departments": ["Electrical"]},
            {"name": "Renewable Energy Systems", "departments": ["Electrical"]},
            {"name": "High Voltage Engineering", "departments": ["Electrical"]},
            {"name": "Electric Drives", "departments": ["Electrical"]},
            {"name": "Protection Systems", "departments": ["Electrical"]},
            
            # Chemical Engineering Department
            {"name": "Chemical Process Design", "departments": ["Chemical"]},
            {"name": "Mass Transfer", "departments": ["Chemical"]},
            {"name": "Chemical Reaction Engineering", "departments": ["Chemical"]},
            {"name": "Process Control", "departments": ["Chemical"]},
            {"name": "Biochemical Engineering", "departments": ["Chemical"]},
            {"name": "Petroleum Engineering", "departments": ["Chemical"]},
            {"name": "Environmental Chemistry", "departments": ["Chemical"]},
        ]
        
        # Insert subjects into database
        result = subjects_col.insert_many(sample_subjects)
        
        return jsonify({
            "message": f"Successfully initialized {len(sample_subjects)} subjects",
            "inserted_count": len(result.inserted_ids)
        }), 201
        
    except Exception as e:
        return jsonify({"message": f"Failed to initialize subjects: {str(e)}"}), 500

@app.route('/api/admin/refresh-subjects', methods=['POST'])
@jwt_required()
def refresh_subjects():
    """Force refresh subjects database with updated data (admin only)"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Clear existing subjects
        deleted_count = subjects_col.delete_many({}).deleted_count
        
        # Re-initialize with updated data (department-only structure)
        sample_subjects = [
            # Computer Science Department
            {"name": "Data Structures", "departments": ["Computer Science"]},
            {"name": "Algorithms", "departments": ["Computer Science"]},
            {"name": "Database Systems", "departments": ["Computer Science"]},
            {"name": "Software Engineering", "departments": ["Computer Science"]},
            {"name": "Operating Systems", "departments": ["Computer Science"]},
            {"name": "Computer Networks", "departments": ["Computer Science"]},
            {"name": "Machine Learning", "departments": ["Computer Science"]},
            {"name": "Web Development", "departments": ["Computer Science"]},
            {"name": "Artificial Intelligence", "departments": ["Computer Science"]},
            {"name": "Computer Graphics", "departments": ["Computer Science"]},
            {"name": "Cybersecurity", "departments": ["Computer Science"]},
            {"name": "Mobile App Development", "departments": ["Computer Science"]},
            
            # Mathematics Department  
            {"name": "Calculus I", "departments": ["Mathematics"]},
            {"name": "Calculus II", "departments": ["Mathematics"]},
            {"name": "Linear Algebra", "departments": ["Mathematics"]},
            {"name": "Differential Equations", "departments": ["Mathematics"]},
            {"name": "Statistics", "departments": ["Mathematics"]},
            {"name": "Discrete Mathematics", "departments": ["Mathematics"]},
            {"name": "Probability Theory", "departments": ["Mathematics"]},
            {"name": "Number Theory", "departments": ["Mathematics"]},
            {"name": "Abstract Algebra", "departments": ["Mathematics"]},
            {"name": "Real Analysis", "departments": ["Mathematics"]},
            
            # Physics Department
            {"name": "Classical Mechanics", "departments": ["Physics"]},
            {"name": "Quantum Physics", "departments": ["Physics"]},
            {"name": "Thermodynamics", "departments": ["Physics"]},
            {"name": "Electromagnetism", "departments": ["Physics"]},
            {"name": "Optics", "departments": ["Physics"]},
            {"name": "Nuclear Physics", "departments": ["Physics"]},
            {"name": "Solid State Physics", "departments": ["Physics"]},
            {"name": "Astrophysics", "departments": ["Physics"]},
            
            # Business Department
            {"name": "Business Management", "departments": ["Business"]},
            {"name": "Marketing", "departments": ["Business"]},
            {"name": "Finance", "departments": ["Business"]},
            {"name": "Economics", "departments": ["Business"]},
            {"name": "Accounting", "departments": ["Business"]},
            {"name": "Human Resources", "departments": ["Business"]},
            {"name": "Operations Management", "departments": ["Business"]},
            {"name": "Business Analytics", "departments": ["Business"]},
            {"name": "Entrepreneurship", "departments": ["Business"]},
            
            # Engineering Department (General)
            {"name": "Engineering Physics", "departments": ["Engineering"]},
            {"name": "Circuit Analysis", "departments": ["Engineering"]},
            {"name": "Digital Logic", "departments": ["Engineering"]},
            {"name": "Control Systems", "departments": ["Engineering"]},
            {"name": "Engineering Mathematics", "departments": ["Engineering"]},
            {"name": "Engineering Drawing", "departments": ["Engineering"]},
            {"name": "Materials Science", "departments": ["Engineering"]},
            
            # ECE (Electronics and Communication Engineering) Department
            {"name": "Electronic Devices", "departments": ["ECE"]},
            {"name": "Analog Electronics", "departments": ["ECE"]},
            {"name": "Digital Electronics", "departments": ["ECE"]},
            {"name": "Communication Systems", "departments": ["ECE"]},
            {"name": "Microprocessors", "departments": ["ECE"]},
            {"name": "VLSI Design", "departments": ["ECE"]},
            {"name": "Signal Processing", "departments": ["ECE"]},
            {"name": "Embedded Systems", "departments": ["ECE"]},
            {"name": "Antenna Theory", "departments": ["ECE"]},
            {"name": "RF Engineering", "departments": ["ECE"]},
            {"name": "Power Electronics", "departments": ["ECE"]},
            {"name": "Digital Signal Processing", "departments": ["ECE"]},
            {"name": "Wireless Communication", "departments": ["ECE"]},
            {"name": "Fiber Optic Communication", "departments": ["ECE"]},
            {"name": "Satellite Communication", "departments": ["ECE"]},
            
            # Mechanical Engineering Department
            {"name": "Thermodynamics", "departments": ["Mechanical"]},
            {"name": "Fluid Mechanics", "departments": ["Mechanical"]},
            {"name": "Machine Design", "departments": ["Mechanical"]},
            {"name": "Heat Transfer", "departments": ["Mechanical"]},
            {"name": "Manufacturing Processes", "departments": ["Mechanical"]},
            {"name": "Strength of Materials", "departments": ["Mechanical"]},
            {"name": "Dynamics of Machinery", "departments": ["Mechanical"]},
            {"name": "Automobile Engineering", "departments": ["Mechanical"]},
            {"name": "Industrial Engineering", "departments": ["Mechanical"]},
            {"name": "Mechanical Vibrations", "departments": ["Mechanical"]},
            
            # Civil Engineering Department  
            {"name": "Structural Analysis", "departments": ["Civil"]},
            {"name": "Concrete Technology", "departments": ["Civil"]},
            {"name": "Geotechnical Engineering", "departments": ["Civil"]},
            {"name": "Transportation Engineering", "departments": ["Civil"]},
            {"name": "Hydraulics", "departments": ["Civil"]},
            {"name": "Environmental Engineering", "departments": ["Civil"]},
            {"name": "Construction Management", "departments": ["Civil"]},
            {"name": "Surveying", "departments": ["Civil"]},
            {"name": "Earthquake Engineering", "departments": ["Civil"]},
            {"name": "Water Resources Engineering", "departments": ["Civil"]},
            
            # Electrical Engineering Department
            {"name": "Electrical Machines", "departments": ["Electrical"]},
            {"name": "Power Systems", "departments": ["Electrical"]},
            {"name": "Electrical Circuits", "departments": ["Electrical"]},
            {"name": "Power Electronics", "departments": ["Electrical"]},
            {"name": "Renewable Energy Systems", "departments": ["Electrical"]},
            {"name": "High Voltage Engineering", "departments": ["Electrical"]},
            {"name": "Electric Drives", "departments": ["Electrical"]},
            {"name": "Protection Systems", "departments": ["Electrical"]},
            
            # Chemical Engineering Department
            {"name": "Chemical Process Design", "departments": ["Chemical"]},
            {"name": "Mass Transfer", "departments": ["Chemical"]},
            {"name": "Chemical Reaction Engineering", "departments": ["Chemical"]},
            {"name": "Process Control", "departments": ["Chemical"]},
            {"name": "Biochemical Engineering", "departments": ["Chemical"]},
            {"name": "Petroleum Engineering", "departments": ["Chemical"]},
            {"name": "Environmental Chemistry", "departments": ["Chemical"]},
        ]
        
        result = subjects_col.insert_many(sample_subjects)
        
        return jsonify({
            "message": f"Successfully refreshed subjects database",
            "deleted_count": deleted_count,
            "inserted_count": len(result.inserted_ids)
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to refresh subjects: {str(e)}"}), 500


@app.route('/api/set-lecture', methods=['POST'])
def set_lecture():
    global current_lecture
    data = request.json
    print(f"DEBUG: Received data: {data}")  # Debug logging
    lecture_number = data.get("lectureNumber")
    lecture_date = data.get("date")
    subject = data.get("subject")  # New: Get subject from request
    print(f"DEBUG: Subject received: '{subject}'")  # Debug logging
    
    if not subject or subject.strip() == "":
        return jsonify({"message": f"Subject is required for lecture. Received: '{subject}' from data: {data}"}), 400
    
    # Check if lecture already exists for this date and subject
    existing_lecture = lectures_col.find_one({
        "lectureNumber": lecture_number,
        "date": lecture_date,
        "subject": subject
    })
    
    if existing_lecture:
        # Set this lecture as the current active lecture
        current_lecture = {
            "_id": str(existing_lecture["_id"]),
            "lectureNumber": lecture_number,
            "date": lecture_date,
            "subject": subject,
            "startTime": existing_lecture.get("startTime", datetime.datetime.now().isoformat()),
            "isActive": True
        }
        
        # Update the lecture to be active
        lectures_col.update_one(
            {"_id": existing_lecture["_id"]},
            {"$set": {"isActive": True}}
        )
        
        return jsonify({
            "message": f"Lecture {lecture_number} for {subject} is ready for attendance on {lecture_date}",
            "lecture": current_lecture
        })
    
    # Create new lecture
    lecture = {
        "lectureNumber": lecture_number,
        "startTime": datetime.datetime.now().isoformat(),
        "date": lecture_date,
        "subject": subject,
        "isActive": True,
        "attendees": []
    }
    
    result = lectures_col.insert_one(lecture)
    current_lecture = {
        "_id": str(result.inserted_id),
        "lectureNumber": lecture_number,
        "date": lecture_date,
        "subject": subject,
        "startTime": lecture["startTime"],
        "isActive": True
    }
    
    return jsonify({
        "message": f"Lecture {lecture_number} for {subject} is ready for attendance on {lecture_date}",
        "lecture": current_lecture
    })

@app.route('/api/start-lecture', methods=['POST'])
def start_lecture():
    global current_lecture
    data = request.json
    lecture_number = data.get("lectureNumber")
    
    if current_lecture:
        return jsonify({"message": "Please end the current lecture first"}), 400
    
    # Create new lecture
    lecture = {
        "lectureNumber": lecture_number,
        "startTime": datetime.datetime.now().isoformat(),
        "date": datetime.datetime.now().strftime("%Y-%m-%d"),
        "isActive": True,
        "attendees": []
    }
    
    result = lectures_col.insert_one(lecture)
    current_lecture = {
        "_id": str(result.inserted_id),
        "lectureNumber": lecture_number,
        "startTime": lecture["startTime"],
        "date": lecture["date"]
    }
    
    print(f"Started lecture {lecture_number}")
    return jsonify({"message": f"Lecture {lecture_number} started successfully", "lecture": current_lecture}), 201

@app.route('/api/end-lecture', methods=['POST'])
def end_lecture():
    global current_lecture
    
    if not current_lecture:
        return jsonify({"message": "No active lecture to end"}), 400
    
    # Update lecture in database
    lectures_col.update_one(
        {"_id": ObjectId(current_lecture["_id"])},
        {
            "$set": {
                "endTime": datetime.datetime.now().isoformat(),
                "isActive": False
            }
        }
    )
    
    lecture_number = current_lecture["lectureNumber"]
    current_lecture = None
    
    print(f"Ended lecture {lecture_number}")
    return jsonify({"message": f"Lecture {lecture_number} ended successfully"}), 200

@app.route('/api/current-lecture', methods=['GET'])
def get_current_lecture():
    return jsonify({"lecture": current_lecture}), 200

@app.route('/api/students/<student_id>', methods=['DELETE'])
@jwt_required()
def delete_student_api(student_id):
    # Verify admin access
    if not verify_admin():
        return jsonify({"message": "Admin access required"}), 403
    """Delete a student and all their attendance records"""
    try:
        # Find the student first
        student = students_col.find_one({"rollNo": student_id})
        if not student:
            return jsonify({"success": False, "error": "Student not found"}), 404
        
        student_name = student.get('name', 'Unknown')
        
        # Delete from students collection
        student_result = students_col.delete_one({"rollNo": student_id})
        if student_result.deleted_count == 0:
            return jsonify({"success": False, "error": "Failed to delete student"}), 500
        
        # Delete all attendance records for this student
        attendance_result = attendance_col.delete_many({"rollNo": student_id})
        
        return jsonify({
            "success": True,
            "message": f"Successfully deleted {student_name}",
            "student_deleted": True,
            "attendance_records_deleted": attendance_result.deleted_count
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/students', methods=['GET'])
def get_students():
    students = list(students_col.find({}, {"_id": 0}))
    return jsonify(students)

@app.route('/api/mark-attendance-safe', methods=['POST'])
def mark_attendance_safe():
    """Safe attendance marking with atomic upsert to prevent duplicates"""
    try:
        data = request.get_json()
        roll_no = data.get('rollNo')
        name = data.get('name')
        status = data.get('status')
        lecture_number = int(data.get('lectureNumber'))  # Ensure integer
        date = data.get('date')
        time = data.get('time')
        method = data.get('method', 'manual')
        
        if not all([roll_no, name, status, lecture_number, date, time]):
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        # Use atomic upsert operation to prevent race conditions
        filter_query = {
            "rollNo": roll_no,
            "lectureNumber": lecture_number,
            "date": date
        }
        
        update_data = {
            "$set": {
                "name": name,
                "status": status,
                "time": time,
                "method": method
            },
            "$setOnInsert": {
                "rollNo": roll_no,
                "lectureNumber": lecture_number,
                "date": date
            }
        }
        
        # Check if record exists first to provide appropriate message
        existing = attendance_col.find_one(filter_query)
        
        result = attendance_col.update_one(
            filter_query,
            update_data,
            upsert=True
        )
        
        if existing:
            return jsonify({
                "success": True,
                "message": f"Updated attendance for {name} to {status}",
                "action": "updated",
                "previousStatus": existing.get("status"),
                "previousMethod": existing.get("method")
            })
        else:
            return jsonify({
                "success": True,
                "message": f"Marked attendance for {name} as {status}",
                "action": "created"
            })
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/mark-attendance-manual', methods=['POST'])
def mark_attendance_manual():
    try:
        data = request.get_json()
        roll_no = data.get('rollNo')
        name = data.get('name')
        status = data.get('status')
        lecture_number = data.get('lectureNumber')
        date = data.get('date')
        time = data.get('time')
        
        if not all([roll_no, name, status, lecture_number, date, time]):
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        # Check if attendance already exists for this student, lecture, date and subject
        subject = current_lecture.get("subject", "Unknown") if current_lecture else "Unknown"
        existing = attendance_col.find_one({
            "rollNo": roll_no,
            "lectureNumber": int(lecture_number),  # Ensure integer comparison
            "date": date,
            "subject": subject
        })
        
        if existing:
            # Always show what exists and ask for confirmation to edit
            existing_method = existing.get("method", "unknown")
            existing_status = existing.get("status")
            existing_time = existing.get("time", "unknown")
            
            # If trying to set the same status, inform user
            if status == existing_status:
                return jsonify({
                    "success": False, 
                    "message": f"Student {name} is already marked {status} for Lecture {lecture_number} at {existing_time} via {existing_method}",
                    "canEdit": True,
                    "currentStatus": existing_status,
                    "currentMethod": existing_method,
                    "currentTime": existing_time,
                    "action": "no_change_needed"
                })
            
            # Allow manual override with clear messaging
            update_message = f"Updated {name} from {existing_status} ({existing_method}) to {status} (manual)"
            
            # UPDATE existing record instead of creating new one
            result = attendance_col.update_one(
                {"_id": existing["_id"]},  # Use exact document ID to avoid any confusion
                {"$set": {
                    "status": status, 
                    "time": time,
                    "subject": subject,
                    "method": "manual"
                }}
            )
            
            if result.modified_count > 0:
                return jsonify({
                    "success": True, 
                    "message": update_message,
                    "action": "updated",
                    "previousStatus": existing_status,
                    "previousMethod": existing_method,
                    "previousTime": existing_time
                })
            else:
                return jsonify({
                    "success": False, 
                    "message": "Failed to update attendance record",
                    "action": "update_failed"
                })
        else:
            # Create new record only if none exists
            # Create new record
            attendance_record = {
                "rollNo": roll_no,
                "name": name,
                "status": status,
                "lectureNumber": int(lecture_number),  # Ensure integer type
                "date": date,
                "subject": subject,
                "time": time,
                "method": "manual"
            }
            attendance_col.insert_one(attendance_record)
            return jsonify({"success": True, "message": "Attendance marked successfully"})
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/lectures', methods=['GET'])
def get_lectures():
    lectures = list(lectures_col.find({}, {"_id": 0}).sort("lectureNumber", -1))
    return jsonify(lectures), 200
def preprocess_image(img):
    """Preprocess image to improve face detection"""
    # Resize if image is too large
    height, width = img.shape[:2]
    if height > 800 or width > 800:
        scale = 800 / max(height, width)
        new_width = int(width * scale)
        new_height = int(height * scale)
        img = cv2.resize(img, (new_width, new_height), interpolation=cv2.INTER_AREA)
        print(f"Resized image to: {new_width}x{new_height}")
    
    # Improve contrast and brightness
    lab = cv2.cvtColor(img, cv2.COLOR_RGB2LAB)
    l_channel, a, b = cv2.split(lab)
    
    # Apply CLAHE to L channel
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    cl = clahe.apply(l_channel)
    
    # Merge channels and convert back to RGB
    enhanced = cv2.merge((cl, a, b))
    enhanced = cv2.cvtColor(enhanced, cv2.COLOR_LAB2RGB)
    
    return enhanced

def get_embedding_from_base64(base64_img):
    try:
        # Decode the base64 image
        img_data = base64.b64decode(base64_img.split(',')[1])
        np_arr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        print(f"Original image shape: {img.shape}, dtype: {img.dtype}")
        
        if img is None:
            raise Exception("Failed to decode image")
        
        # Convert BGR to RGB for DeepFace
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Try a simpler approach first - just use OpenCV backend without enforcement
        try:
            print("Trying OpenCV backend without strict enforcement...")
            embedding = DeepFace.represent(
                img_rgb, 
                model_name='Facenet',
                detector_backend='opencv',
                enforce_detection=False
            )[0]['embedding']
            print("SUCCESS: Face embedding extracted with OpenCV")
            return embedding
        except Exception as e:
            print(f"OpenCV approach failed: {e}")
        
        # Try with MTCNN backend
        try:
            print("Trying MTCNN backend...")
            embedding = DeepFace.represent(
                img_rgb, 
                model_name='Facenet',
                detector_backend='mtcnn',
                enforce_detection=False
            )[0]['embedding']
            print("SUCCESS: Face embedding extracted with MTCNN")
            return embedding
        except Exception as e:
            print(f"MTCNN approach failed: {e}")
        
        # Last resort: try with basic settings
        try:
            print("Trying basic DeepFace settings...")
            embedding = DeepFace.represent(img_rgb, model_name='Facenet')[0]['embedding']
            print("SUCCESS: Face embedding extracted with basic settings")
            return embedding
        except Exception as e:
            print(f"Basic approach also failed: {e}")
            raise Exception(f"All face detection methods failed. Last error: {e}")
        
    except Exception as e:
        print(f"Error in get_embedding_from_base64: {e}")
        raise

@app.route('/api/enroll', methods=['POST'])
@jwt_required()
def enroll_student():
    # Verify admin access
    if not verify_admin():
        return jsonify({"message": "Admin access required"}), 403
    data = request.json
    print(f"Received enrollment data: {data}")
    name = data.get("name")
    rollNo = data.get("rollNo")
    batch = data.get("batch")
    group = data.get("group")
    images = data.get("images", [])
    print(f"Name: {name}, RollNo: {rollNo}, Batch: {batch}, Group: {group}, Number of images: {len(images)}")
    
    # Debug: Check current database state
    all_students = list(students_col.find({}))
    print(f"DEBUG: Current students in database: {len(all_students)}")
    for s in all_students:
        print(f"DEBUG: - {s.get('name')} (Roll: {s.get('rollNo')})")
    
    # Check if roll number already exists
    existing_student = students_col.find_one({"rollNo": rollNo})
    if existing_student:
        print(f"DEBUG: Found existing student: {existing_student.get('name')} with roll {rollNo}")
        return jsonify({"message": f"Student with roll number {rollNo} already exists!"}), 400
    
    # Extract embeddings from new images
    new_embeddings = []
    for idx, img_b64 in enumerate(images):
        try:
            print(f"DEBUG: Processing image {idx+1}/{len(images)} for {name}")
            emb = get_embedding_from_base64(img_b64)
            new_embeddings.append(emb)
            print(f"DEBUG: Image {idx+1}: Embedding extracted successfully. Shape: {len(emb)}")
        except Exception as e:
            print(f"ERROR: Image {idx+1}: Failed to extract embedding. Error: {e}")
            continue
    
    if not new_embeddings:
        print(f"ERROR: No valid faces detected in any image for {name}")
        return jsonify({"message": "No valid face detected in any image! Please ensure the face is clearly visible and well-lit."}), 400
    
    print(f"DEBUG: Successfully extracted {len(new_embeddings)} embeddings for {name}")
    
    # Check for face similarity with existing students
    all_students = list(students_col.find({}))
    print(f"DEBUG: Checking face similarity against {len(all_students)} existing students")
    
    # Use stricter threshold for enrollment to prevent false duplicates
    # Lower threshold = stricter matching (only very similar faces are considered duplicates)
    enrollment_threshold = 8.0  # Much stricter than attendance matching (15)
    
    duplicate_found = False
    min_distance = float('inf')
    closest_student = None
    
    for student in all_students:
        student_name = student.get('name', 'Unknown')
        existing_embeddings = student.get('embeddings', [])
        
        for i, existing_emb in enumerate(existing_embeddings):
            for j, new_emb in enumerate(new_embeddings):
                dist = np.linalg.norm(np.array(new_emb) - np.array(existing_emb))
                
                # Track minimum distance for debugging
                if dist < min_distance:
                    min_distance = dist
                    closest_student = student_name
                
                if dist < enrollment_threshold:
                    print(f"ERROR: Duplicate face detected! {name} vs {student_name}: distance {dist:.2f} < {enrollment_threshold}")
                    return jsonify({
                        "message": f"Face already exists! Very similar to student {student_name} (Roll: {student['rollNo']}) - Distance: {dist:.2f}"
                    }), 400
    
    print(f"DEBUG: Face similarity check passed. Closest match: {closest_student} (distance: {min_distance:.2f})")
    
    # If no duplicates found, proceed with enrollment
    print(f"DEBUG: No face duplicates found. Proceeding with enrollment for {name}")
    student = {
        "name": name,
        "rollNo": rollNo,
        "batch": batch if batch else None,
        "group": group if group else None,
        "embeddings": new_embeddings,
        "images": images
    }
    try:
        result = students_col.insert_one(student)
        print(f"Student inserted with _id: {result.inserted_id}")
    except Exception as e:
        print(f"Failed to insert student: {e}")
        return jsonify({"message": f"Failed to enroll student: {e}"}), 500
    return jsonify({"message": "Student enrolled successfully", "embeddings_saved": len(new_embeddings)}), 201


@app.route('/api/mark-attendance', methods=['POST'])
def mark_attendance():
    global current_lecture
    
    if not current_lecture:
        return jsonify({"message": "No active lecture. Please ask teacher to start a lecture first."}), 400
    
    data = request.json
    image_b64 = data.get("image")
    if not image_b64:
        return jsonify({"message": "No image provided"}), 400

    try:
        captured_embedding = get_embedding_from_base64(image_b64)
    except Exception as e:
        print(f"Face detection error: {e}")
        return jsonify({"message": f"Face not detected: {e}"}), 400

    # Find all students and compare embeddings
    students = list(students_col.find({}))
    min_dist = float('inf')
    matched_student = None
    for student in students:
        for emb in student.get('embeddings', []):
            dist = np.linalg.norm(np.array(captured_embedding) - np.array(emb))
            if dist < min_dist:
                min_dist = dist
                matched_student = student

    # Set a threshold for matching (tune as needed)
    threshold = 15  # Increased threshold for more tolerant matching
    if matched_student and min_dist < threshold:
        roll_no = matched_student['rollNo']
        
        # Check if student is already present in this lecture (either method)
        existing_attendance = attendance_col.find_one({
            "rollNo": roll_no,
            "lectureNumber": current_lecture["lectureNumber"],
            "date": current_lecture["date"],
            "subject": current_lecture.get("subject", "Unknown")
        })
        
        if existing_attendance:
            existing_status = existing_attendance.get("status")
            existing_method = existing_attendance.get("method", "manual")
            
            if existing_status == "Present":
                return jsonify({
                    "message": f"{matched_student['name']} is already marked present for Lecture {current_lecture['lectureNumber']} via {existing_method}"
                }), 200
            elif existing_status == "Absent" and existing_method == "manual":
                # Don't override manual absent marking with face recognition
                return jsonify({
                    "message": f"{matched_student['name']} is manually marked absent for Lecture {current_lecture['lectureNumber']}. Cannot override with face recognition."
                }), 400
            else:
                # Update existing attendance from Absent to Present (only if not manually set to absent)
                attendance_col.update_one(
                    {"_id": existing_attendance["_id"]},
                    {"$set": {
                        "status": "Present",
                        "time": datetime.datetime.now().strftime("%H:%M:%S"),
                        "subject": current_lecture.get("subject", "Unknown"),
                        "method": "face_recognition"
                    }}
                )
                return jsonify({
                    "message": f"Attendance updated to Present for {matched_student['name']} (Roll: {roll_no}) - Lecture {current_lecture['lectureNumber']}"
                }), 200
        
        # Mark attendance
        attendance_record = {
            "rollNo": roll_no,
            "name": matched_student['name'],
            "lectureNumber": current_lecture["lectureNumber"],
            "date": current_lecture["date"],
            "subject": current_lecture.get("subject", "Unknown"),
            "time": datetime.datetime.now().strftime("%H:%M:%S"),
            "status": "Present",
            "method": "face_recognition"
        }
        attendance_col.insert_one(attendance_record)
        
        # Add to lecture attendees
        lectures_col.update_one(
            {"_id": ObjectId(current_lecture["_id"])},
            {"$addToSet": {"attendees": roll_no}}
        )
        
        return jsonify({"message": f"Attendance marked for {matched_student['name']} (Roll: {roll_no}) - Lecture {current_lecture['lectureNumber']}"}), 200
    else:
        return jsonify({"message": "No matching student found"}), 404

@app.route('/api/attendance-records', methods=['GET'])
def get_attendance_records():
    lecture_number = request.args.get('lectureNumber')
    date_filter = request.args.get('date')
    query = {}
    
    if lecture_number:
        query["lectureNumber"] = int(lecture_number)
    if date_filter:
        query["date"] = date_filter
    
    records = list(attendance_col.find(query, {"_id": 0}).sort("lectureNumber", -1))
    return jsonify(records)

@app.route('/api/cleanup-attendance', methods=['DELETE'])
def cleanup_attendance():
    try:
        # Delete records with missing or invalid data
        result = attendance_col.delete_many({
            "$or": [
                {"name": {"$exists": False}},
                {"name": ""},
                {"name": None},
                {"lectureNumber": {"$exists": False}},
                {"lectureNumber": None},
                {"rollNo": {"$exists": False}},
                {"rollNo": ""},
                {"rollNo": None}
            ]
        })
        
        return jsonify({
            "success": True,
            "message": f"Cleaned up {result.deleted_count} invalid attendance records"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/delete-attendance-record', methods=['DELETE'])
def delete_attendance_record():
    try:
        data = request.get_json()
        roll_no = data.get('rollNo')
        date = data.get('date')
        time = data.get('time')
        
        if not all([roll_no, date, time]):
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        result = attendance_col.delete_one({
            "rollNo": roll_no,
            "date": date,
            "time": time
        })
        
        if result.deleted_count > 0:
            return jsonify({"success": True, "message": "Record deleted successfully"})
        else:
            return jsonify({"success": False, "error": "Record not found"}), 404
            
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    try:
        # Get total number of lectures
        total_lectures = lectures_col.count_documents({})
        
        # Get total attendance records
        total_attendance = attendance_col.count_documents({})
        
        # Get total enrolled students
        total_students = students_col.count_documents({})
        
        # Calculate average attendance per lecture
        if total_lectures > 0:
            average_attendance = total_attendance / total_lectures
        else:
            average_attendance = 0
        
        # Get today's attendance
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        today_attendance = attendance_col.count_documents({"date": today})
        
        # Get this week's attendance
        week_ago = (datetime.datetime.now() - datetime.timedelta(days=7)).strftime('%Y-%m-%d')
        week_attendance = attendance_col.count_documents({
            "date": {"$gte": week_ago}
        })
        
        # Get active lectures count
        active_lectures = lectures_col.count_documents({"isActive": True})
        
        return jsonify({
            "totalLectures": total_lectures,
            "totalAttendance": total_attendance,
            "totalStudents": total_students,
            "averageAttendance": round(average_attendance, 1),
            "todayAttendance": today_attendance,
            "weekAttendance": week_attendance,
            "activeLectures": active_lectures
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/manual-attendance', methods=['POST'])
def mark_manual_attendance():
    try:
        data = request.json
        student_ids = data.get('studentIds', [])
        status = data.get('status', 'Present')  # Present, Absent, Late
        lecture_number = data.get('lectureNumber')
        date = data.get('date')
        
        if not student_ids or not lecture_number or not date:
            return jsonify({"success": False, "error": "Missing required fields"}), 400
        
        results = []
        
        for student_id in student_ids:
            # Get student details
            student = students_col.find_one({"rollNo": student_id})
            if not student:
                results.append({"studentId": student_id, "success": False, "error": "Student not found"})
                continue
            
            # Check if attendance already exists
            existing_attendance = attendance_col.find_one({
                "rollNo": student_id,
                "date": date,
                "lectureNumber": int(lecture_number)  # Ensure integer comparison
            })
            
            if existing_attendance:
                existing_status = existing_attendance.get('status')
                existing_method = existing_attendance.get('method', 'manual')
                
                # Allow manual override but provide informative feedback
                if status == existing_status:
                    results.append({
                        "studentId": student_id, 
                        "success": False, 
                        "action": "already_marked", 
                        "message": f"Already marked {status} via {existing_method}",
                        "currentStatus": existing_status,
                        "currentMethod": existing_method
                    })
                else:
                    # Update existing attendance with manual override
                    update_message = f"Updated from {existing_status} ({existing_method}) to {status} (manual)"
                    attendance_col.update_one(
                        {"_id": existing_attendance["_id"]},
                        {"$set": {
                            "status": status,
                            "time": datetime.datetime.now().strftime('%H:%M:%S'),
                            "method": "manual"
                        }}
                    )
                    results.append({
                        "studentId": student_id, 
                        "success": True, 
                        "action": "updated",
                        "message": update_message,
                        "previousStatus": existing_status,
                        "previousMethod": existing_method
                    })
            else:
                # Create new attendance record
                attendance_record = {
                    "name": student.get("name", "Unknown"),
                    "rollNo": student_id,
                    "lectureNumber": int(lecture_number),  # Ensure integer type
                    "date": date,
                    "time": datetime.datetime.now().strftime('%H:%M:%S'),
                    "status": status,
                    "method": "manual"
                }
                attendance_col.insert_one(attendance_record)
                results.append({"studentId": student_id, "success": True, "action": "created"})
        
        return jsonify({
            "success": True,
            "message": f"Manual attendance marked for {len([r for r in results if r['success']])} students",
            "results": results
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/bulk-attendance', methods=['POST'])
def bulk_attendance_upload():
    try:
        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Get lecture details from form data
        lecture_number = request.form.get('lectureNumber')
        date = request.form.get('date')
        
        if not lecture_number or not date:
            return jsonify({"error": "Missing lectureNumber or date"}), 400
        
        # Read Excel file
        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(BytesIO(file.read()))
            
            # Validate required columns
            required_columns = ['Roll No.', 'Name', 'Attendance']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    "error": f"Missing required columns: {', '.join(missing_columns)}. Required columns are: {', '.join(required_columns)}"
                }), 400
            
            # Clean and validate data
            df = df.dropna(subset=['Roll No.', 'Attendance'])  # Remove rows with missing roll no or attendance
            df['Roll No.'] = df['Roll No.'].astype(str)  # Ensure roll numbers are strings
            df['Attendance'] = df['Attendance'].str.upper()  # Convert to uppercase
            
            # Validate attendance values
            valid_attendance_values = ['P', 'A']
            invalid_attendance = df[~df['Attendance'].isin(valid_attendance_values)]
            if not invalid_attendance.empty:
                return jsonify({
                    "error": f"Invalid attendance values found. Use 'P' for Present or 'A' for Absent. Invalid rows: {invalid_attendance.index.tolist()}"
                }), 400
            
        except Exception as e:
            return jsonify({"error": f"Error reading Excel file: {str(e)}"}), 400
        
        # Process attendance records
        results = []
        processed_count = 0
        
        for index, row in df.iterrows():
            try:
                roll_no = str(row['Roll No.']).strip()
                student_name = str(row['Name']).strip() if pd.notna(row['Name']) else "Unknown"
                attendance_status = 'present' if row['Attendance'] == 'P' else 'absent'
                
                # Verify student exists in database
                student = students_col.find_one({"rollNo": roll_no})
                if not student:
                    results.append({
                        "rollNo": roll_no,
                        "name": student_name,
                        "success": False,
                        "message": f"Student with Roll No. {roll_no} not found in database"
                    })
                    continue
                
                # Check if attendance already exists
                existing_attendance = attendance_col.find_one({
                    "rollNo": roll_no,
                    "lectureNumber": int(lecture_number),
                    "date": date
                })
                
                if existing_attendance:
                    # Update existing attendance
                    attendance_col.update_one(
                        {"_id": existing_attendance["_id"]},
                        {"$set": {
                            "status": attendance_status,
                            "time": datetime.datetime.now().strftime('%H:%M:%S'),
                            "method": "bulk_upload"
                        }}
                    )
                    results.append({
                        "rollNo": roll_no,
                        "name": student_name,
                        "success": True,
                        "message": f"Updated attendance to {attendance_status}"
                    })
                else:
                    # Create new attendance record
                    attendance_record = {
                        "name": student.get("name", student_name),
                        "rollNo": roll_no,
                        "lectureNumber": int(lecture_number),
                        "date": date,
                        "time": datetime.datetime.now().strftime('%H:%M:%S'),
                        "status": attendance_status,
                        "method": "bulk_upload"
                    }
                    attendance_col.insert_one(attendance_record)
                    results.append({
                        "rollNo": roll_no,
                        "name": student_name,
                        "success": True,
                        "message": f"Marked {attendance_status}"
                    })
                
                processed_count += 1
                
            except Exception as e:
                results.append({
                    "rollNo": roll_no if 'roll_no' in locals() else f"Row {index + 1}",
                    "name": student_name if 'student_name' in locals() else "Unknown",
                    "success": False,
                    "message": f"Error processing row: {str(e)}"
                })
        
        successful_count = len([r for r in results if r['success']])
        
        return jsonify({
            "success": True,
            "message": f"Excel attendance processed. {successful_count} out of {len(results)} records processed successfully.",
            "processed": processed_count,
            "successful": successful_count,
            "total_rows": len(df),
            "results": results
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/students-attendance-status', methods=['GET'])
def get_students_attendance_status():
    try:
        lecture_number = request.args.get('lectureNumber')
        date = request.args.get('date')
        
        if not lecture_number or not date:
            return jsonify({"error": "Missing lectureNumber or date"}), 400
        
        # Get all students
        students = list(students_col.find({}, {"_id": 0, "name": 1, "rollNo": 1, "email": 1}))
        
        # Get attendance for this lecture and date
        attendance_records = list(attendance_col.find({
            "lectureNumber": lecture_number,
            "date": date
        }, {"_id": 0, "rollNo": 1, "status": 1, "method": 1}))
        
        # Create a map of attendance status
        attendance_map = {record['rollNo']: record for record in attendance_records}
        
        # Combine student data with attendance status
        students_with_status = []
        for student in students:
            attendance = attendance_map.get(student['rollNo'])
            students_with_status.append({
                "name": student['name'],
                "rollNo": student['rollNo'],
                "email": student.get('email', ''),
                "status": attendance['status'] if attendance else 'Not Marked',
                "method": attendance.get('method', '') if attendance else ''
            })
        
        return jsonify(students_with_status)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/debug-student-records', methods=['GET'])
def debug_student_records():
    """Debug specific student records"""
    try:
        roll_no = request.args.get('rollNo', '2310993858')
        lecture_number = request.args.get('lectureNumber', 15)
        date = request.args.get('date', '2025-09-08')
        
        # Search with both integer and string lecture numbers
        records1 = list(attendance_col.find({
            "rollNo": roll_no,
            "lectureNumber": int(lecture_number),
            "date": date
        }))
        
        records2 = list(attendance_col.find({
            "rollNo": roll_no,
            "lectureNumber": str(lecture_number),
            "date": date
        }))
        
        # Also search without lecture number constraint
        all_records = list(attendance_col.find({
            "rollNo": roll_no,
            "date": date
        }))
        
        # Convert ObjectId to string for JSON serialization
        for records in [records1, records2, all_records]:
            for record in records:
                record['_id'] = str(record['_id'])
            
        return jsonify({
            "success": True,
            "with_int_lecture": {"count": len(records1), "records": records1},
            "with_str_lecture": {"count": len(records2), "records": records2},
            "all_for_date": {"count": len(all_records), "records": all_records},
            "query": {
                "rollNo": roll_no,
                "lectureNumber": lecture_number,
                "date": date
            }
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/check-duplicates', methods=['GET'])
def check_duplicates():
    """Check for duplicate attendance records"""
    try:
        # Simple way to find potential duplicates
        pipeline = [
            {
                "$group": {
                    "_id": {
                        "rollNo": "$rollNo",
                        "lectureNumber": "$lectureNumber", 
                        "date": "$date"
                    },
                    "count": {"$sum": 1},
                    "docs": {"$push": {
                        "id": "$_id",
                        "name": "$name",
                        "status": "$status",
                        "time": "$time",
                        "method": "$method"
                    }}
                }
            },
            {
                "$match": {
                    "count": {"$gt": 1}
                }
            }
        ]
        
        duplicates = list(attendance_col.aggregate(pipeline))
        
        return jsonify({
            "success": True,
            "duplicates_found": len(duplicates),
            "duplicates": duplicates
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/cleanup-duplicate-attendance', methods=['POST'])
def cleanup_duplicate_attendance():
    """Remove duplicate attendance records for the same student, lecture, and date"""
    try:
        # First, normalize lectureNumber data types - convert all to integers
        attendance_col.update_many(
            {"lectureNumber": {"$type": "string"}},
            [{"$set": {"lectureNumber": {"$toInt": "$lectureNumber"}}}]
        )
        
        # Now find duplicates with normalized data
        pipeline = [
            {
                "$group": {
                    "_id": {
                        "rollNo": "$rollNo",
                        "lectureNumber": "$lectureNumber", 
                        "date": "$date"
                    },
                    "count": {"$sum": 1},
                    "docs": {"$push": "$$ROOT"}
                }
            },
            {
                "$match": {
                    "count": {"$gt": 1}
                }
            }
        ]
        
        duplicates = list(attendance_col.aggregate(pipeline))
        removed_count = 0
        
        for duplicate_group in duplicates:
            docs = duplicate_group["docs"]
            
            # Priority: Keep Present over Absent, face_recognition over manual
            best_record = None
            records_to_delete = []
            
            for doc in docs:
                if best_record is None:
                    best_record = doc
                else:
                    # Choose better record based on priority
                    current_better = False
                    
                    # Present status is better than Absent
                    if doc.get("status") == "Present" and best_record.get("status") != "Present":
                        current_better = True
                    # If both have same status, prefer face_recognition over manual
                    elif (doc.get("status") == best_record.get("status") and 
                          doc.get("method") == "face_recognition" and 
                          best_record.get("method") == "manual"):
                        current_better = True
                    
                    if current_better:
                        records_to_delete.append(best_record["_id"])
                        best_record = doc
                    else:
                        records_to_delete.append(doc["_id"])
            
            # Delete inferior records
            for record_id in records_to_delete:
                attendance_col.delete_one({"_id": record_id})
                removed_count += 1
        
        return jsonify({
            "success": True,
            "message": f"Cleaned up {removed_count} duplicate attendance records",
            "duplicates_found": len(duplicates),
            "data_normalization": "Converted all lectureNumber fields to integers"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/download-students-excel', methods=['GET'])
@jwt_required()
def download_students_excel():
    """
    Download student database as Excel file containing names and roll numbers
    """
    try:
        print("DEBUG: Excel download endpoint called")
        current_user = get_jwt_identity()
        print(f"DEBUG: Current user from JWT: {current_user}")
        
        # Get all students from the database
        students = list(students_col.find({}))
        print(f"DEBUG: Found {len(students)} students in database")
        
        if not students:
            print("DEBUG: No students found")
            return jsonify({"error": "No students found in database"}), 404
        
        # Create a new workbook and select the active worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Student Database"
        print("DEBUG: Created workbook and worksheet")
        
        # Set up headers
        headers = ["S.No", "Name", "Roll Number", "Batch", "Group/Section"]
        worksheet.append(headers)
        print("DEBUG: Added headers")
        
        # Style the headers
        for cell in worksheet[1]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = cell.font.copy(color="FFFFFF")
        print("DEBUG: Styled headers")
        
        # Add student data
        for idx, student in enumerate(students, start=1):
            row_data = [
                idx,  # Serial number
                student.get('name', 'N/A'),
                student.get('rollNo', 'N/A'),
                student.get('batch', 'N/A'),
                student.get('group', 'N/A')
            ]
            worksheet.append(row_data)
        print(f"DEBUG: Added {len(students)} rows of student data")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        print("DEBUG: Adjusted column widths")
        
        # Save workbook to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        print("DEBUG: Saved workbook to buffer")
        
        # Generate filename with current timestamp
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"student_database_{timestamp}.xlsx"
        print(f"DEBUG: Generated filename: {filename}")
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"ERROR: Exception in Excel download: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate Excel file: {str(e)}"}), 500

# ============== DATA MANAGEMENT ENDPOINTS ==============

# Helper function to generate unique student username
def generate_unique_student_username(name):
    """
    Generate a unique student username in format: First letter (capitalized) + 5-digit random number
    Example: A59382
    """
    if not name:
        raise ValueError("Name cannot be empty")
    
    # Get first letter and capitalize it
    first_letter = name.strip()[0].upper()
    
    # Keep generating until we find a unique username
    max_attempts = 100
    attempts = 0
    
    while attempts < max_attempts:
        # Generate random 5-digit number (10000-99999)
        random_number = random.randint(10000, 99999)
        username = f"{first_letter}{random_number}"
        
        # Check if username already exists in database
        if not students_col.find_one({"username": username}):
            return username
        
        attempts += 1
    
    raise Exception("Could not generate unique student username after maximum attempts")

def generate_unique_student_password():
    """
    Generate a unique 8-character random alphanumeric password
    Uses Python's secrets module for cryptographically strong random generation
    """
    import secrets
    import string
    
    max_attempts = 100
    attempts = 0
    
    while attempts < max_attempts:
        # Generate 8-character alphanumeric password
        alphabet = string.ascii_letters + string.digits
        password = ''.join(secrets.choice(alphabet) for _ in range(8))
        
        # Check if password already exists (check hashed version)
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        if not students_col.find_one({"password_hash": password_hash}):
            return password
        
        attempts += 1
    
    raise Exception("Could not generate unique student password after maximum attempts")

def send_student_credentials_email(student_name, username, password, email):
    """
    Send student credentials via email
    """
    try:
        sender_email = "karanjulka2512@gmail.com"
        sender_password = "mdaw pumn vupx zxdg"
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        
        print(f"📧 Sending student credentials to: {email}")
        
        # Create message
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = email
        message["Subject"] = "Your Student Login Credentials – Smart AI Attendance System"
        
        # Email body
        body = f"""Hello {student_name},

Welcome to the Smart AI-Powered Attendance System!

Here are your login credentials:

Username: {username}
Password: {password}

Please log in using these credentials to access the student portal.

IMPORTANT: Please change your password after your first login for security purposes.

Best regards,
Smart Attendance System Admin
"""
        
        message.attach(MIMEText(body, "plain"))
        
        # Connect to server and send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        text = message.as_string()
        server.sendmail(sender_email, email, text)
        server.quit()
        print(f"✅ Student credentials email sent successfully to: {email}")
        
        return True
    except Exception as e:
        print(f"❌ Student email sending error: {str(e)}")
        return False

@app.route('/api/admin/students', methods=['GET'])
@jwt_required()
def get_all_students():
    """Get all students with filtering and sorting"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all students (excluding sensitive data)
        students = list(students_col.find({}, {
            "_id": 0,
            "password_hash": 0,
            "embeddings": 0,
            "images": 0
        }))
        
        return jsonify({"students": students}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to fetch students: {str(e)}"}), 500

@app.route('/api/admin/teachers', methods=['GET'])
@jwt_required()
def get_all_teachers():
    """Get all teachers with filtering and sorting"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all teachers (excluding sensitive data, but include _id for deletion)
        teachers = list(teachers_col.find({}, {
            "password_hash": 0
        }))
        
        # Convert ObjectId to string for JSON serialization
        for teacher in teachers:
            if '_id' in teacher:
                teacher['_id'] = str(teacher['_id'])
        
        return jsonify({"teachers": teachers}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to fetch teachers: {str(e)}"}), 500

@app.route('/api/admin/students/<student_id>', methods=['PUT'])
@jwt_required()
def update_student(student_id):
    """Update student details and generate credentials if needed"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        
        # Find existing student
        existing_student = students_col.find_one({"rollNo": student_id})
        if not existing_student:
            return jsonify({"message": "Student not found"}), 404
        
        # Check if student previously had no password (first-time credential generation)
        needs_credentials = not existing_student.get("password_hash")
        
        # Prepare update data
        update_data = {}
        if data.get('name'):
            update_data['name'] = data['name']
        if data.get('email'):
            update_data['email'] = data['email']
        if data.get('department'):
            update_data['department'] = data['department']
        if data.get('course'):
            update_data['course'] = data['course']
        if 'phone' in data:
            update_data['phone'] = data['phone']
        if 'batch' in data:
            # Validate batch if provided
            batch = data['batch']
            if batch and (not batch.isdigit() or len(batch) != 4):
                return jsonify({"message": "Batch must be a 4-digit year (e.g., 2023)"}), 400
            update_data['batch'] = batch if batch else None
        if 'group' in data:
            # Validate group if provided
            group = data['group']
            if group is not None and not str(group).strip():
                return jsonify({"message": "Group cannot be empty"}), 400
            update_data['group'] = group if group else None
        
        # Generate credentials if this is first time or explicitly requested
        credentials_generated = False
        if needs_credentials:
            username = generate_unique_student_username(update_data.get('name', existing_student['name']))
            password = generate_unique_student_password()
            password_hash = hashlib.sha256(password.encode()).hexdigest()
            
            update_data['username'] = username
            update_data['password_hash'] = password_hash
            update_data['email_sent'] = False
            
            credentials_generated = True
        
        # Update student record
        result = students_col.update_one(
            {"rollNo": student_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0 and not credentials_generated:
            return jsonify({"message": "No changes made"}), 200
        
        # Send email if credentials were generated
        email_sent = False
        if credentials_generated:
            student_email = update_data.get('email', existing_student.get('email'))
            student_name = update_data.get('name', existing_student.get('name'))
            
            if student_email:
                email_sent = send_student_credentials_email(
                    student_name,
                    username,
                    password,
                    student_email
                )
                
                if email_sent:
                    students_col.update_one(
                        {"rollNo": student_id},
                        {"$set": {"email_sent": True}}
                    )
        
        response_message = "Student updated successfully"
        if credentials_generated:
            response_message += f". Login credentials generated (Username: {username})"
            if email_sent:
                response_message += " and sent via email"
            else:
                response_message += " but email failed to send"
        
        return jsonify({
            "message": response_message,
            "credentials_generated": credentials_generated,
            "email_sent": email_sent
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to update student: {str(e)}"}), 500

@app.route('/api/admin/students/<student_id>', methods=['DELETE'])
@jwt_required()
def delete_student(student_id):
    """Delete student and all related attendance records"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Find student first
        student = students_col.find_one({"rollNo": student_id})
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        student_name = student.get('name', 'Unknown')
        
        # Delete student record
        student_result = students_col.delete_one({"rollNo": student_id})
        
        # Cascade delete: Remove all attendance records for this student
        attendance_result = attendance_col.delete_many({"rollNo": student_id})
        
        return jsonify({
            "success": True,
            "message": f"Successfully deleted {student_name} and {attendance_result.deleted_count} attendance record(s)",
            "student_deleted": True,
            "attendance_records_deleted": attendance_result.deleted_count
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to delete student: {str(e)}"}), 500

@app.route('/api/admin/teachers/<teacher_id>', methods=['PUT'])
@jwt_required()
def update_teacher(teacher_id):
    """Update teacher details (supports both username and _id)"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        
        # Try to find teacher by username first, then by _id
        existing_teacher = teachers_col.find_one({"username": teacher_id})
        
        # If not found by username, try by _id (for legacy teachers)
        if not existing_teacher:
            try:
                from bson import ObjectId
                existing_teacher = teachers_col.find_one({"_id": ObjectId(teacher_id)})
            except:
                pass
        
        if not existing_teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Prepare update data
        update_data = {}
        if data.get('name'):
            update_data['name'] = data['name']
        if data.get('email'):
            update_data['email'] = data['email']
        if data.get('department'):
            update_data['department'] = data['department']
        if 'contact_number' in data:
            update_data['contact_number'] = data['contact_number']
        
        # Update teacher record by _id (most reliable)
        result = teachers_col.update_one(
            {"_id": existing_teacher['_id']},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            return jsonify({"message": "No changes made"}), 200
        
        return jsonify({"message": "Teacher updated successfully"}), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to update teacher: {str(e)}"}), 500

@app.route('/api/admin/teachers/<teacher_id>', methods=['DELETE'])
@jwt_required()
def delete_teacher(teacher_id):
    """Delete teacher and related assignments (supports both username and _id)"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Try to find teacher by username first, then by _id
        teacher = teachers_col.find_one({"username": teacher_id})
        
        # If not found by username, try by _id (for legacy teachers without username)
        if not teacher:
            try:
                from bson import ObjectId
                teacher = teachers_col.find_one({"_id": ObjectId(teacher_id)})
            except:
                pass
        
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        teacher_name = teacher.get('name', 'Unknown')
        teacher_username = teacher.get('username', None)
        teacher_obj_id = teacher.get('_id')
        
        # Check if teacher is currently a mentor - PREVENT DELETION
        mentor_assignment = mentor_assignments_col.find_one({"mentor_teacher_id": str(teacher_obj_id)})
        if mentor_assignment:
            return jsonify({
                "message": f"Cannot delete {teacher_name}. They are currently mentoring Batch {mentor_assignment['batch']} - Group {mentor_assignment['group']}. Please remove the mentor assignment first.",
                "is_mentor": True,
                "batch": mentor_assignment['batch'],
                "group": mentor_assignment['group']
            }), 409
        
        # Delete teacher record by _id (most reliable)
        teacher_result = teachers_col.delete_one({"_id": teacher_obj_id})
        
        # Cascade delete: Remove all assignments for this teacher
        # Use username if available, otherwise use email as fallback
        if teacher_username:
            assignments_result = teacher_assignments_col.delete_many({"teacher_username": teacher_username})
        else:
            # For legacy teachers, try to find assignments by email or _id
            teacher_email = teacher.get('email', '')
            assignments_result = teacher_assignments_col.delete_many({
                "$or": [
                    {"teacher_email": teacher_email},
                    {"teacher_id": str(teacher_obj_id)}
                ]
            })
        
        return jsonify({
            "success": True,
            "message": f"Successfully deleted {teacher_name} and {assignments_result.deleted_count} assignment(s)",
            "teacher_deleted": True,
            "assignments_deleted": assignments_result.deleted_count
        }), 200
        
    except Exception as e:
        return jsonify({"message": f"Failed to delete teacher: {str(e)}"}), 500

@app.route('/api/admin/students/export', methods=['GET'])
@jwt_required()
def export_students_excel():
    """Export all students to formatted Excel file"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all students
        students = list(students_col.find({}, {
            "_id": 0,
            "password_hash": 0,
            "embeddings": 0,
            "images": 0
        }))
        
        if not students:
            return jsonify({"message": "No students found"}), 404
        
        # Create DataFrame
        df = pd.DataFrame(students)
        
        # Reorder columns for better readability - include username, phone, batch, and group
        column_order = ['name', 'email', 'department', 'rollNo', 'batch', 'group', 'username', 'phone', 'course', 'email_sent', 'created_at']
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # Rename columns for Excel
        df.rename(columns={
            'name': 'Name',
            'email': 'Email',
            'department': 'Department',
            'rollNo': 'Enrollment Number',
            'batch': 'Batch (Year)',
            'group': 'Group/Section',
            'username': 'Username (Login)',
            'phone': 'Phone Number',
            'course': 'Course',
            'email_sent': 'Credentials Sent',
            'created_at': 'Enrollment Date'
        }, inplace=True)
        
        # Create Excel file with formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Students', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Students']
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"students_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"message": f"Failed to export students: {str(e)}"}), 500

@app.route('/api/admin/teachers/export', methods=['GET'])
@jwt_required()
def export_teachers_excel():
    """Export all teachers to formatted Excel file"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all teachers
        teachers = list(teachers_col.find({}, {
            "_id": 0,
            "password_hash": 0
        }))
        
        if not teachers:
            return jsonify({"message": "No teachers found"}), 404
        
        # Create DataFrame
        df = pd.DataFrame(teachers)
        
        # Reorder columns for better readability - include username and phone
        column_order = ['name', 'username', 'email', 'department', 'contact_number', 'created_at']
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        # Rename columns for Excel
        df.rename(columns={
            'name': 'Name',
            'username': 'Username (Login)',
            'email': 'Email',
            'department': 'Department',
            'contact_number': 'Phone Number',
            'created_at': 'Added On'
        }, inplace=True)
        
        # Create Excel file with formatting
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Teachers', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Teachers']
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"teachers_export_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({"message": f"Failed to export teachers: {str(e)}"}), 500

# Modify existing add-student endpoint to include credential generation
@app.route('/api/admin/add-student-with-credentials', methods=['POST'])
@jwt_required()
def add_student_with_credentials():
    """Add new student with automatic credential generation and email"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        name = data.get('name')
        roll_no = data.get('roll_no')
        email = data.get('email')
        course = data.get('course')
        department = data.get('department')
        images = data.get('images', [])
        
        if not all([name, roll_no, email, course, department]):
            return jsonify({"message": "Name, roll number, email, course, and department are required"}), 400
        
        # Check if student already exists
        if students_col.find_one({"rollNo": roll_no}):
            return jsonify({"message": "Student with this roll number already exists"}), 400
        
        if students_col.find_one({"email": email}):
            return jsonify({"message": "Student with this email already exists"}), 400
        
        # Generate credentials
        username = generate_unique_student_username(name)
        password = generate_unique_student_password()
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Process face recognition if images provided
        embeddings = []
        if images:
            for idx, img_b64 in enumerate(images):
                try:
                    emb = get_embedding_from_base64(img_b64)
                    embeddings.append(emb)
                except Exception as e:
                    print(f"Failed to process image {idx+1}: {e}")
                    continue
        
        # Create student record with credentials
        student_data = {
            "name": name,
            "rollNo": roll_no,
            "email": email,
            "course": course,
            "department": department,
            "batch": batch if batch else None,
            "group": group if group else None,
            "username": username,
            "password_hash": password_hash,
            "email_sent": False,
            "embeddings": embeddings,
            "images": images,
            "created_at": datetime.datetime.now()
        }
        
        result = students_col.insert_one(student_data)
        
        # Send credentials email
        email_sent = send_student_credentials_email(name, username, password, email)
        
        if email_sent:
            students_col.update_one(
                {"_id": result.inserted_id},
                {"$set": {"email_sent": True}}
            )
        
        return jsonify({
            "message": "Student added successfully",
            "student_id": str(result.inserted_id),
            "username": username,
            "password": password,
            "email_sent": email_sent,
            "embeddings_processed": len(embeddings)
        }), 201
        
    except Exception as e:
        return jsonify({"message": f"Failed to add student: {str(e)}"}), 500

# ============== END DATA MANAGEMENT ENDPOINTS ==============

# ============== STUDENT PORTAL ENDPOINTS ==============

@app.route('/api/student/login', methods=['POST'])
def student_login():
    """Student login endpoint - authenticate using username and password"""
    try:
        username = request.json.get('username')
        password = request.json.get('password')
        
        if not username or not password:
            return jsonify({"message": "Username and password required"}), 400
        
        # Find student in database by username
        student = students_col.find_one({"username": username})
        
        if not student:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Check if student has a password set
        if not student.get('password_hash'):
            return jsonify({"message": "No password set. Please contact administrator."}), 401
        
        # Verify password using SHA256 hashing
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if student.get('password_hash') != password_hash:
            return jsonify({"message": "Invalid credentials"}), 401
        
        # Update last login timestamp
        students_col.update_one(
            {"username": username},
            {"$set": {"last_login": datetime.datetime.now()}}
        )
        
        # Create JWT token with student role
        access_token = create_access_token(
            identity=username, 
            additional_claims={"role": "student", "rollNo": student.get('rollNo')}
        )
        
        return jsonify({
            "token": access_token,
            "user": {
                "username": student['username'],
                "name": student['name'],
                "email": student['email'],
                "rollNo": student.get('rollNo'),
                "department": student.get('department'),
                "role": "student"
            }
        }), 200
        
    except Exception as e:
        print(f"Student login error: {e}")
        return jsonify({"message": "Login failed"}), 500

@app.route('/api/student/profile', methods=['GET'])
@jwt_required()
def get_student_profile():
    """Get student profile information"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a student
        if claims.get('role') != 'student':
            return jsonify({"message": "Access denied. Students only."}), 403
        
        # Find student by username
        student = students_col.find_one(
            {"username": current_user},
            {
                "_id": 0,
                "password_hash": 0,
                "embeddings": 0,
                "images": 0
            }
        )
        
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        return jsonify({
            "profile": {
                "name": student.get('name'),
                "username": student.get('username'),
                "email": student.get('email'),
                "rollNo": student.get('rollNo'),
                "department": student.get('department'),
                "course": student.get('course'),
                "phone": student.get('phone'),
                "batch": student.get('batch'),
                "group": student.get('group'),
                "mentor_name": student.get('mentor_name'),
                "mentor_email": student.get('mentor_email'),
                "mentor_id": student.get('mentor_id'),
                "created_at": student.get('created_at')
            }
        }), 200
        
    except Exception as e:
        print(f"Get student profile error: {e}")
        return jsonify({"message": "Failed to fetch profile"}), 500

@app.route('/api/student/attendance', methods=['GET'])
@jwt_required()
def get_student_attendance():
    """Get subject-wise attendance for logged-in student with accurate calculations"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a student
        if claims.get('role') != 'student':
            return jsonify({"message": "Access denied. Students only."}), 403
        
        roll_no = claims.get('rollNo')
        
        if not roll_no:
            return jsonify({"message": "Roll number not found in token"}), 400
        
        # Find student to get department
        student = students_col.find_one({"rollNo": roll_no})
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        student_department = student.get('department')
        student_name = student.get('name')
        
        # Get all attendance records for this student
        all_attendance = list(attendance_col.find({"rollNo": roll_no}))
        
        if not all_attendance:
            # Return empty attendance with student info
            return jsonify({
                "overall": {
                    "total_classes": 0,
                    "present": 0,
                    "absent": 0,
                    "percentage": 0.0
                },
                "subjects": [],
                "student_name": student_name,
                "roll_no": roll_no,
                "department": student_department
            }), 200
        
        # Calculate OVERALL attendance correctly
        total_classes = len(all_attendance)
        present_count = 0
        absent_count = 0
        dl_count = 0
        ml_count = 0
        
        for record in all_attendance:
            status = record.get('status', '').lower()
            
            if status == 'present':
                present_count += 1
            elif status in ['absent', 'absent (not marked)']:
                absent_count += 1
            elif status == 'dl':
                dl_count += 1
            elif status == 'ml':
                ml_count += 1
        
        # Calculate overall percentage - ML counts as present
        if total_classes == 0:
            overall_percentage = 0.0
        else:
            # ML is considered present
            effective_present = present_count + ml_count
            overall_percentage = (effective_present / total_classes) * 100
            overall_percentage = min(round(overall_percentage, 2), 100.0)  # Cap at 100%
        
        # Get teacher assignments for subject-wise breakdown
        teacher_assignments = list(teacher_assignments_col.find({"department": student_department}))
        
        subjects_data = []
        
        if teacher_assignments:
            # Create subject-wise attendance breakdown
            for assignment in teacher_assignments:
                subject_name = assignment.get('subject')
                teacher_username = assignment.get('teacher_username')
                teacher_name = assignment.get('teacher_name', 'Unknown Teacher')
                
                # Skip if no subject name
                if not subject_name:
                    continue
                
                # Filter attendance records for this subject/teacher if teacher_id exists
                # Since current records don't have subject_id, we'll show overall stats per subject
                # In a real implementation, records would have subject_id or teacher_id
                subject_records = all_attendance  # Would filter by subject_id if available
                
                subject_total = len(subject_records)
                subject_present = sum(1 for r in subject_records if r.get('status', '').lower() == 'present')
                subject_absent = sum(1 for r in subject_records if r.get('status', '').lower() in ['absent', 'absent (not marked)'])
                subject_dl = sum(1 for r in subject_records if r.get('status', '').lower() == 'dl')
                subject_ml = sum(1 for r in subject_records if r.get('status', '').lower() == 'ml')
                
                # Calculate subject percentage - ML counts as present
                if subject_total == 0:
                    subject_percentage = 0.0
                else:
                    # ML is considered present for attendance percentage
                    effective_present = subject_present + subject_ml
                    subject_percentage = (effective_present / subject_total) * 100
                    subject_percentage = min(round(subject_percentage, 2), 100.0)  # Cap at 100%
                
                subjects_data.append({
                    "subject_name": subject_name,
                    "teacher_name": teacher_name,
                    "delivered": subject_total,
                    "attended": subject_present,
                    "dl": subject_dl,
                    "ml": subject_ml,
                    "overall_percentage": subject_percentage
                })
        else:
            # No subjects found - create a single "Overall Attendance" entry
            # ML counts as present
            effective_present = present_count + ml_count
            if total_classes == 0:
                calc_percentage = 0.0
            else:
                calc_percentage = (effective_present / total_classes) * 100
                calc_percentage = min(round(calc_percentage, 2), 100.0)
            
            subjects_data.append({
                "subject_name": "Overall Attendance",
                "teacher_name": "All Teachers",
                "delivered": total_classes,
                "attended": present_count,
                "dl": dl_count,
                "ml": ml_count,
                "overall_percentage": calc_percentage
            })
        
        # Return structured response
        return jsonify({
            "overall": {
                "total_classes": total_classes,
                "present": present_count,
                "absent": absent_count,
                "dl": dl_count,
                "ml": ml_count,
                "percentage": overall_percentage
            },
            "subjects": subjects_data,
            "student_name": student_name,
            "roll_no": roll_no,
            "department": student_department
        }), 200
        
    except Exception as e:
        print(f"Get student attendance error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"message": f"Failed to fetch attendance: {str(e)}"}), 500

# ============== END STUDENT PORTAL ENDPOINTS ==============

# ============== MENTOR ASSIGNMENT ENDPOINTS ==============

@app.route('/api/admin/assign-mentor', methods=['POST'])
@jwt_required()
def assign_mentor():
    """Assign a mentor teacher to a batch+group combination"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        batch = data.get('batch')
        group = data.get('group')
        teacher_id = data.get('teacher_id')
        
        if not all([batch, group, teacher_id]):
            return jsonify({"message": "Batch, group, and teacher_id are required"}), 400
        
        # Validate batch format (must be 4 digits)
        if not batch.isdigit() or len(batch) != 4:
            return jsonify({"message": "Invalid batch format. Must be a 4-digit year"}), 400
        
        # Validate group
        if not group.strip():
            return jsonify({"message": "Group cannot be empty"}), 400
        
        # Check if teacher exists
        teacher = teachers_col.find_one({"_id": ObjectId(teacher_id)})
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Check if this batch+group has any students
        students_count = students_col.count_documents({"batch": batch, "group": group})
        if students_count == 0:
            return jsonify({"message": f"No students found for Batch {batch} - Group {group}"}), 404
        
        # Check if this teacher is already mentoring another batch+group
        existing_mentor_assignment = mentor_assignments_col.find_one({"mentor_teacher_id": teacher_id})
        if existing_mentor_assignment:
            return jsonify({
                "message": f"Teacher {teacher.get('name')} is already mentoring Batch {existing_mentor_assignment['batch']} - Group {existing_mentor_assignment['group']}"
            }), 409
        
        # Check if this batch+group already has a mentor
        existing_batch_group = mentor_assignments_col.find_one({"batch": batch, "group": group})
        
        mentor_data = {
            "batch": batch,
            "group": group,
            "mentor_teacher_id": teacher_id,
            "mentor_teacher_name": teacher.get('name'),
            "mentor_email": teacher.get('email'),
            "assigned_at": datetime.datetime.now(),
            "assigned_by": get_jwt_identity()
        }
        
        if existing_batch_group:
            # Update existing assignment
            mentor_assignments_col.update_one(
                {"batch": batch, "group": group},
                {"$set": mentor_data}
            )
            action = "updated"
        else:
            # Create new assignment
            mentor_assignments_col.insert_one(mentor_data)
            action = "assigned"
        
        # Update all students in this batch+group with mentor information
        update_result = students_col.update_many(
            {"batch": batch, "group": group},
            {"$set": {
                "mentor_id": teacher_id,
                "mentor_name": teacher.get('name'),
                "mentor_email": teacher.get('email')
            }}
        )
        
        return jsonify({
            "message": f"Mentor {action} successfully",
            "mentor_name": teacher.get('name'),
            "batch": batch,
            "group": group,
            "students_updated": update_result.modified_count
        }), 200
        
    except Exception as e:
        print(f"Assign mentor error: {e}")
        return jsonify({"message": f"Failed to assign mentor: {str(e)}"}), 500

@app.route('/api/admin/mentor-assignments', methods=['GET'])
@jwt_required()
def get_mentor_assignments():
    """Get all mentor assignments"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all mentor assignments
        assignments = list(mentor_assignments_col.find({}, {"_id": 0}))
        
        # Enrich with student count for each assignment
        for assignment in assignments:
            student_count = students_col.count_documents({
                "batch": assignment['batch'],
                "group": assignment['group']
            })
            assignment['student_count'] = student_count
        
        return jsonify({"assignments": assignments}), 200
        
    except Exception as e:
        print(f"Get mentor assignments error: {e}")
        return jsonify({"message": f"Failed to fetch mentor assignments: {str(e)}"}), 500

@app.route('/api/admin/unassigned-groups', methods=['GET'])
@jwt_required()
def get_unassigned_groups():
    """Get all batch+group combinations that don't have a mentor assigned"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all unique batch+group combinations from students
        pipeline = [
            {"$match": {"batch": {"$ne": None}, "group": {"$ne": None}}},
            {"$group": {
                "_id": {"batch": "$batch", "group": "$group"},
                "student_count": {"$sum": 1}
            }},
            {"$project": {
                "_id": 0,
                "batch": "$_id.batch",
                "group": "$_id.group",
                "student_count": 1
            }}
        ]
        
        all_groups = list(students_col.aggregate(pipeline))
        
        # Get all assigned batch+group combinations
        assigned_groups = set()
        for assignment in mentor_assignments_col.find({}, {"batch": 1, "group": 1}):
            assigned_groups.add((assignment['batch'], assignment['group']))
        
        # Filter out assigned groups
        unassigned = [
            group for group in all_groups
            if (group['batch'], group['group']) not in assigned_groups
        ]
        
        return jsonify({"unassigned_groups": unassigned}), 200
        
    except Exception as e:
        print(f"Get unassigned groups error: {e}")
        return jsonify({"message": f"Failed to fetch unassigned groups: {str(e)}"}), 500

@app.route('/api/admin/available-mentors', methods=['GET'])
@jwt_required()
def get_available_mentors():
    """Get teachers who are not currently assigned as mentors"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        # Get all teachers
        all_teachers = list(teachers_col.find({}, {
            "_id": 1,
            "name": 1,
            "email": 1,
            "department": 1
        }))
        
        # Get assigned mentor teacher IDs
        assigned_teacher_ids = set()
        for assignment in mentor_assignments_col.find({}, {"mentor_teacher_id": 1}):
            assigned_teacher_ids.add(assignment['mentor_teacher_id'])
        
        # Filter available teachers
        available = [
            {
                "teacher_id": str(teacher['_id']),
                "name": teacher.get('name'),
                "email": teacher.get('email'),
                "department": teacher.get('department')
            }
            for teacher in all_teachers
            if str(teacher['_id']) not in assigned_teacher_ids
        ]
        
        return jsonify({"available_mentors": available}), 200
        
    except Exception as e:
        print(f"Get available mentors error: {e}")
        return jsonify({"message": f"Failed to fetch available mentors: {str(e)}"}), 500

@app.route('/api/admin/remove-mentor', methods=['DELETE'])
@jwt_required()
def remove_mentor_assignment():
    """Remove mentor assignment for a batch+group"""
    try:
        # Verify admin access
        if not verify_admin():
            return jsonify({"message": "Admin access required"}), 403
        
        data = request.json
        batch = data.get('batch')
        group = data.get('group')
        
        if not all([batch, group]):
            return jsonify({"message": "Batch and group are required"}), 400
        
        # Remove mentor assignment
        result = mentor_assignments_col.delete_one({"batch": batch, "group": group})
        
        if result.deleted_count == 0:
            return jsonify({"message": "No mentor assignment found for this batch+group"}), 404
        
        # Remove mentor info from students
        students_col.update_many(
            {"batch": batch, "group": group},
            {"$unset": {"mentor_id": "", "mentor_name": "", "mentor_email": ""}}
        )
        
        return jsonify({"message": "Mentor assignment removed successfully"}), 200
        
    except Exception as e:
        print(f"Remove mentor error: {e}")
        return jsonify({"message": f"Failed to remove mentor: {str(e)}"}), 500

@app.route('/api/student/mentor', methods=['GET'])
@jwt_required()
def get_student_mentor():
    """Get mentor information for the logged-in student"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a student
        if claims.get('role') != 'student':
            return jsonify({"message": "Access denied. Students only."}), 403
        
        # Find student by username
        student = students_col.find_one(
            {"username": current_user},
            {"mentor_id": 1, "mentor_name": 1, "mentor_email": 1, "batch": 1, "group": 1}
        )
        
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        if not student.get('mentor_name'):
            return jsonify({
                "has_mentor": False,
                "message": "No mentor assigned to your batch and group yet"
            }), 200
        
        return jsonify({
            "has_mentor": True,
            "mentor_name": student.get('mentor_name'),
            "mentor_email": student.get('mentor_email'),
            "batch": student.get('batch'),
            "group": student.get('group')
        }), 200
        
    except Exception as e:
        print(f"Get student mentor error: {e}")
        return jsonify({"message": f"Failed to fetch mentor: {str(e)}"}), 500

@app.route('/api/teacher/mentor-group', methods=['GET'])
@jwt_required()
def get_teacher_mentor_group():
    """Get the batch+group that the logged-in teacher is mentoring"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a teacher
        if claims.get('role') != 'teacher':
            return jsonify({"message": "Access denied. Teachers only."}), 403
        
        # Find teacher by username
        teacher = teachers_col.find_one({"username": current_user}, {"_id": 1})
        
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Find mentor assignment
        assignment = mentor_assignments_col.find_one(
            {"mentor_teacher_id": str(teacher['_id'])},
            {"_id": 0, "batch": 1, "group": 1, "assigned_at": 1}
        )
        
        if not assignment:
            return jsonify({
                "is_mentor": False,
                "message": "You are not assigned as a mentor to any batch and group"
            }), 200
        
        # Count students in this batch+group
        student_count = students_col.count_documents({
            "batch": assignment['batch'],
            "group": assignment['group']
        })
        
        return jsonify({
            "is_mentor": True,
            "batch": assignment['batch'],
            "group": assignment['group'],
            "student_count": student_count,
            "assigned_at": assignment.get('assigned_at')
        }), 200
        
    except Exception as e:
        print(f"Get teacher mentor group error: {e}")
        return jsonify({"message": f"Failed to fetch mentor group: {str(e)}"}), 500

# ============== END MENTOR ASSIGNMENT ENDPOINTS ==============

# ============== MEDICAL LEAVE (ML) ENDPOINTS ==============

@app.route('/api/student/apply_ml', methods=['POST'])
@jwt_required()
def apply_medical_leave():
    """Student applies for medical leave with date range and PDF proof"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a student
        if claims.get('role') != 'student':
            return jsonify({"message": "Access denied. Students only."}), 403
        
        # Get student details
        student = students_col.find_one({"username": current_user})
        
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        # Check if student has a mentor assigned
        if not student.get('mentor_id'):
            return jsonify({"message": "You don't have a mentor assigned. Please contact admin."}), 400
        
        # Get form data
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({"message": "Start date and end date are required"}), 400
        
        # Validate dates
        try:
            start = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.datetime.strptime(end_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({"message": "Invalid date format. Use YYYY-MM-DD"}), 400
        
        if start > end:
            return jsonify({"message": "Start date cannot be after end date"}), 400
        
        # Validate date range (not more than 30 days in past or future)
        today = datetime.datetime.now()
        if start < today - datetime.timedelta(days=30):
            return jsonify({"message": "Cannot apply for leave more than 30 days in the past"}), 400
        
        if end > today + datetime.timedelta(days=90):
            return jsonify({"message": "Cannot apply for leave more than 90 days in advance"}), 400
        
        # Handle file upload
        if 'proof' not in request.files:
            return jsonify({"message": "Medical proof PDF is required"}), 400
        
        file = request.files['proof']
        
        if file.filename == '':
            return jsonify({"message": "No file selected"}), 400
        
        # Validate file type (PDF only)
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"message": "Only PDF files are allowed"}), 400
        
        # Generate unique filename
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{student.get('rollNo')}_{timestamp}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Save file
        file.save(filepath)
        
        # Create ML request document
        ml_request = {
            "student_rollNo": student.get('rollNo'),
            "student_name": student.get('name'),
            "student_email": student.get('email'),
            "batch": student.get('batch'),
            "group": student.get('group'),
            "mentor_id": student.get('mentor_id'),
            "mentor_name": student.get('mentor_name'),
            "mentor_email": student.get('mentor_email'),
            "start_date": start_date,
            "end_date": end_date,
            "proof_filename": filename,
            "proof_url": f"/uploads/ml_proofs/{filename}",
            "status": "PENDING",
            "applied_at": datetime.datetime.utcnow(),
            "updated_at": datetime.datetime.utcnow()
        }
        
        result = medical_leaves_col.insert_one(ml_request)
        
        return jsonify({
            "message": "Medical leave application submitted successfully",
            "request_id": str(result.inserted_id),
            "status": "PENDING"
        }), 201
        
    except Exception as e:
        print(f"Apply ML error: {e}")
        return jsonify({"message": f"Failed to submit application: {str(e)}"}), 500

@app.route('/api/student/ml_requests', methods=['GET'])
@jwt_required()
def get_student_ml_requests():
    """Get all medical leave requests for logged-in student"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a student
        if claims.get('role') != 'student':
            return jsonify({"message": "Access denied. Students only."}), 403
        
        # Get student rollNo
        student = students_col.find_one({"username": current_user}, {"rollNo": 1})
        
        if not student:
            return jsonify({"message": "Student not found"}), 404
        
        # Get all ML requests for this student
        requests = list(medical_leaves_col.find(
            {"student_rollNo": student.get('rollNo')},
            {"_id": 1, "start_date": 1, "end_date": 1, "status": 1, "applied_at": 1, 
             "updated_at": 1, "proof_filename": 1, "mentor_name": 1}
        ).sort("applied_at", -1))
        
        # Convert ObjectId to string
        for req in requests:
            req['_id'] = str(req['_id'])
            req['applied_at'] = req['applied_at'].isoformat() if req.get('applied_at') else None
            req['updated_at'] = req['updated_at'].isoformat() if req.get('updated_at') else None
        
        return jsonify({
            "requests": requests,
            "total": len(requests)
        }), 200
        
    except Exception as e:
        print(f"Get student ML requests error: {e}")
        return jsonify({"message": f"Failed to fetch requests: {str(e)}"}), 500

@app.route('/api/teacher/ml_requests', methods=['GET'])
@jwt_required()
def get_mentor_ml_requests():
    """Get medical leave requests for students mentored by logged-in teacher"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a teacher
        if claims.get('role') != 'teacher':
            return jsonify({"message": "Access denied. Teachers only."}), 403
        
        # Find teacher by username
        teacher = teachers_col.find_one({"username": current_user}, {"_id": 1, "name": 1})
        
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Check if teacher is a mentor
        mentor_assignment = mentor_assignments_col.find_one(
            {"mentor_teacher_id": str(teacher['_id'])}
        )
        
        if not mentor_assignment:
            return jsonify({
                "is_mentor": False,
                "message": "You are not assigned as a mentor",
                "requests": []
            }), 200
        
        # Get all ML requests for students mentored by this teacher
        requests = list(medical_leaves_col.find(
            {"mentor_id": str(teacher['_id'])},
            {"_id": 1, "student_rollNo": 1, "student_name": 1, "student_email": 1,
             "batch": 1, "group": 1, "start_date": 1, "end_date": 1, "status": 1,
             "applied_at": 1, "updated_at": 1, "proof_filename": 1, "proof_url": 1}
        ).sort("applied_at", -1))
        
        # Convert ObjectId to string and format dates
        for req in requests:
            req['_id'] = str(req['_id'])
            req['applied_at'] = req['applied_at'].isoformat() if req.get('applied_at') else None
            req['updated_at'] = req['updated_at'].isoformat() if req.get('updated_at') else None
        
        return jsonify({
            "is_mentor": True,
            "batch": mentor_assignment.get('batch'),
            "group": mentor_assignment.get('group'),
            "requests": requests,
            "total": len(requests),
            "pending": len([r for r in requests if r['status'] == 'PENDING'])
        }), 200
        
    except Exception as e:
        print(f"Get mentor ML requests error: {e}")
        return jsonify({"message": f"Failed to fetch requests: {str(e)}"}), 500

@app.route('/api/teacher/approve_ml', methods=['POST'])
@jwt_required()
def approve_reject_ml():
    """Mentor approves or rejects a medical leave request"""
    try:
        # Get current user identity from JWT
        current_user = get_jwt_identity()
        claims = get_jwt()
        
        # Verify this is a teacher
        if claims.get('role') != 'teacher':
            return jsonify({"message": "Access denied. Teachers only."}), 403
        
        # Find teacher by username
        teacher = teachers_col.find_one({"username": current_user}, {"_id": 1, "name": 1})
        
        if not teacher:
            return jsonify({"message": "Teacher not found"}), 404
        
        # Get request data
        data = request.get_json()
        request_id = data.get('request_id')
        decision = data.get('decision')  # "APPROVED" or "REJECTED"
        
        if not request_id or not decision:
            return jsonify({"message": "Request ID and decision are required"}), 400
        
        if decision not in ['APPROVED', 'REJECTED']:
            return jsonify({"message": "Decision must be APPROVED or REJECTED"}), 400
        
        # Get the ML request
        try:
            ml_request = medical_leaves_col.find_one({"_id": ObjectId(request_id)})
        except:
            return jsonify({"message": "Invalid request ID"}), 400
        
        if not ml_request:
            return jsonify({"message": "ML request not found"}), 404
        
        # Verify this teacher is the mentor for this student
        if ml_request.get('mentor_id') != str(teacher['_id']):
            return jsonify({"message": "You are not the mentor for this student"}), 403
        
        # Check if already processed
        if ml_request.get('status') != 'PENDING':
            return jsonify({"message": f"Request already {ml_request.get('status')}"}), 400
        
        # Update ML request status
        medical_leaves_col.update_one(
            {"_id": ObjectId(request_id)},
            {
                "$set": {
                    "status": decision,
                    "updated_at": datetime.datetime.utcnow(),
                    "processed_by": teacher.get('name')
                }
            }
        )
        
        # If APPROVED, update attendance records
        if decision == "APPROVED":
            start_date = datetime.datetime.strptime(ml_request['start_date'], '%Y-%m-%d')
            end_date = datetime.datetime.strptime(ml_request['end_date'], '%Y-%m-%d')
            student_rollNo = ml_request['student_rollNo']
            
            # Generate list of dates in the range
            date_list = []
            current = start_date
            while current <= end_date:
                date_list.append(current.strftime('%Y-%m-%d'))
                current += datetime.timedelta(days=1)
            
            # CORRECT LOGIC: Update only ABSENT lecture entries to ML
            # Count how many actual lecture entries were updated
            ml_lectures_updated = 0
            total_absent_found = 0
            total_lectures_checked = 0
            
            for date_str in date_list:
                # Find ALL attendance records for this date and student
                # There may be MULTIPLE lectures per day
                attendance_records = list(attendance_col.find({
                    "rollNo": student_rollNo,
                    "date": date_str
                }))
                
                total_lectures_checked += len(attendance_records)
                
                # Update each ABSENT lecture to ML
                for record in attendance_records:
                    if record.get('status') == 'Absent':
                        total_absent_found += 1
                        # Update this specific lecture entry from Absent to ML
                        result = attendance_col.update_one(
                            {"_id": record['_id']},
                            {
                                "$set": {
                                    "status": "ML",
                                    "method": "ml_approved",
                                    "updated_at": datetime.datetime.utcnow()
                                }
                            }
                        )
                        if result.modified_count > 0:
                            ml_lectures_updated += 1
                    # If status is "Present", leave it as Present (do nothing)
                    # ML should NOT override Present attendance
            
            # Log the results for debugging
            print(f"ML Approval for {student_rollNo}:")
            print(f"  - Date range: {ml_request['start_date']} to {ml_request['end_date']}")
            print(f"  - Total dates: {len(date_list)}")
            print(f"  - Total lecture entries checked: {total_lectures_checked}")
            print(f"  - Absent lectures found: {total_absent_found}")
            print(f"  - ML lectures updated: {ml_lectures_updated}")
            
            return jsonify({
                "message": f"Medical leave approved successfully",
                "status": "APPROVED",
                "ml_lectures_updated": ml_lectures_updated,
                "total_absent_found": total_absent_found,
                "dates_covered": len(date_list),
                "total_lectures_checked": total_lectures_checked,
                "details": f"Updated {ml_lectures_updated} absent lecture(s) to ML across {len(date_list)} day(s)"
            }), 200
        else:
            return jsonify({
                "message": "Medical leave rejected",
                "status": "REJECTED"
            }), 200
        
    except Exception as e:
        print(f"Approve/Reject ML error: {e}")
        return jsonify({"message": f"Failed to process request: {str(e)}"}), 500

@app.route('/uploads/ml_proofs/<filename>', methods=['GET'])
@jwt_required()
def download_ml_proof(filename):
    """Download or view ML proof PDF"""
    try:
        claims = get_jwt()
        role = claims.get('role')
        
        # Only students and teachers can access
        if role not in ['student', 'teacher']:
            return jsonify({"message": "Access denied"}), 403
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath):
            return jsonify({"message": "File not found"}), 404
        
        return send_file(filepath, mimetype='application/pdf')
        
    except Exception as e:
        print(f"Download ML proof error: {e}")
        return jsonify({"message": f"Failed to download file: {str(e)}"}), 500

# ============== END MEDICAL LEAVE ENDPOINTS ==============

if __name__ == '__main__':

    app.run(debug=True, host='127.0.0.1', port=5001)
